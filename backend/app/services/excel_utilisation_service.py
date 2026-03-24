"""
Service for parsing and storing the Excel utilisation report.

Sheet layout:
  - Row 4 (index 3): header with datetime month columns
  - Row 5+ (index 4+): data rows
  - Col 2: Name, Col 3: Status, Col 4: Business Unit, Col 5: Function,
    Col 6: Practice, Col 7: Sub-Practice
  - Month columns: header cell is a datetime(YYYY, M, 1) object
  - Cell values: 0.0-1.0 fractions where 0 = fully utilised and 1 = fully available

Classification:
  availability 0-30% -> fully_billed
  availability 31-70% -> partially_billed
  availability 71-100% -> bench
"""

import calendar
import io
import re
import secrets
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Callable, Optional
from urllib.parse import unquote

import openpyxl
from bson import ObjectId

from app.config import settings
from app.models.department import Department
from app.models.employee import Employee
from app.models.employee_project import EmployeeProject
from app.models.employee_skill import EmployeeSkill
from app.models.excel_upload_log import ExcelUploadLog
from app.models.excel_utilisation import ExcelUtilisationReport
from app.models.location import Location
from app.models.project_allocation import ProjectAllocation
from app.models.project import Project
from app.models.reporting_relationship import ReportingRelationship
from app.services.billable_hours_service import (
    get_effective_timesheet_entries,
    summarise_hours_by_employee,
    summarise_hours_by_employee_project,
)

YTPL_SHEET_NAMES = ["Availability Report - YTPL", "Availability Report - NGL"]
EXCEL_TIMESHEET_PROJECT_ID = "excel-utilisation-report"
EXCEL_TIMESHEET_PROJECT_NAME = "Excel Utilisation Report"
EXCEL_EMPLOYEE_PREFIX = "excel:"
DEFAULT_PLANNED_WORKED_DAYS = 20.0


def _classify(availability_pct: float) -> str:
    if availability_pct <= 30:
        return "fully_billed"
    if availability_pct <= 70:
        return "partially_billed"
    return "bench"


def _parse_remarks_availability(text) -> Optional[float]:
    """
    Parse availability fraction (0.0–1.0) from a free-text Remarks cell.
    Returns None if the employee should be skipped (e.g. maternity/accident).
    """
    if text is None:
        return None
    t = str(text).strip().lower()
    if not t:
        return None
    # Skip — not available
    if any(kw in t for kw in ("not available", "maternity", "accident")):
        return None
    # Fully billed
    if "fully billed" in t or "100% billed" in t:
        return 0.0
    # Fully available / bench
    if "fully available" in t or "100% available" in t:
        return 1.0
    if any(kw in t for kw in ("cannot be billed", "corporate", "released", "awaiting", "will be released", "blocked")):
        return 1.0
    # "X% available for billing"
    m = re.search(r"(\d+(?:\.\d+)?)\s*%\s*available", t)
    if m:
        return float(m.group(1)) / 100.0
    # "X% billed" → availability = 1 - X%
    m = re.search(r"(\d+(?:\.\d+)?)\s*%\s*billed", t)
    if m:
        return 1.0 - float(m.group(1)) / 100.0
    return None


def _normalise_text(value) -> str:
    return str(value).strip() if value is not None else ""


def _normalise_name(value: str) -> str:
    cleaned = value.lower().strip()
    cleaned = re.sub(r"\s*[-(]?\s*(onsite|offshore)\s*[)]?\s*$", "", cleaned)
    cleaned = re.sub(r"[^a-z0-9]+", "", cleaned)
    return cleaned


def _normalise_employee_key(employee_name: str, employee_id: Optional[str]) -> str:
    if employee_id:
        return f"id:{employee_id}"
    return f"name:{_normalise_name(employee_name)}"


def _excel_employee_ref(employee_name: str) -> str:
    return f"{EXCEL_EMPLOYEE_PREFIX}{employee_name}"


def _decode_excel_employee_ref(employee_ref: str) -> str:
    return unquote(employee_ref[len(EXCEL_EMPLOYEE_PREFIX):]) if employee_ref.startswith(EXCEL_EMPLOYEE_PREFIX) else employee_ref


def _has_excel_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _prefer_excel_value(excel_value, fallback_value, default):
    if _has_excel_value(excel_value):
        return excel_value
    if _has_excel_value(fallback_value):
        return fallback_value
    return default


def _header_period(cell_value) -> Optional[str]:
    if isinstance(cell_value, datetime):
        return f"{cell_value.year:04d}-{cell_value.month:02d}"
    if isinstance(cell_value, date):
        return f"{cell_value.year:04d}-{cell_value.month:02d}"
    return None


def _dedupe_latest_rows(rows: list[ExcelUtilisationReport]) -> list[ExcelUtilisationReport]:
    latest: dict[str, ExcelUtilisationReport] = {}
    for row in rows:
        key = _normalise_employee_key(row.employee_name, row.employee_id)
        if key not in latest or row.uploaded_at > latest[key].uploaded_at:
            latest[key] = row
    return list(latest.values())


async def _get_latest_excel_period(branch_location_id: str) -> Optional[str]:
    latest_period_doc = await ExcelUtilisationReport.find(
        ExcelUtilisationReport.branch_location_id == branch_location_id,
    ).sort(-ExcelUtilisationReport.period).limit(1).to_list()
    return latest_period_doc[0].period if latest_period_doc else None


def _select_sheet(workbook):
    exact_names = {name.strip().lower(): name for name in workbook.sheetnames}
    for preferred_name in YTPL_SHEET_NAMES:
        actual_name = exact_names.get(preferred_name.strip().lower())
        if actual_name:
            return workbook[actual_name]

    for name in workbook.sheetnames:
        if "ytpl" in name.strip().lower():
            return workbook[name]

    for name in workbook.sheetnames:
        lowered = name.strip().lower()
        if "availability" in lowered and "ngl" not in lowered:
            return workbook[name]

    return workbook.active


def _parse_ytpl_sheet(ws) -> tuple[dict[int, str], int]:
    """Find the header row and build a col_index -> YYYY-MM map for month columns."""
    rows = list(ws.iter_rows(values_only=True))
    for row_index, row in enumerate(rows[:15]):
        month_cols: dict[int, str] = {}
        for col_index, cell in enumerate(row):
            period = _header_period(cell)
            if period:
                month_cols[col_index] = period
        if month_cols:
            return month_cols, row_index
    return {}, -1


def _working_days_in_period(period: str) -> int:
    year = int(period[:4])
    month = int(period[5:7])
    _, days_in_month = calendar.monthrange(year, month)
    return sum(1 for day in range(1, days_in_month + 1) if date(year, month, day).weekday() < 5)


def _capacity_hours_for_period(period: str) -> float:
    return round(_working_days_in_period(period) * 8.0, 2)


def _period_date(period: str) -> str:
    return f"{period}-01"


def _period_end_date(period: str) -> str:
    try:
        year, month = int(period[:4]), int(period[5:7])
        last_day = calendar.monthrange(year, month)[1]
        return f"{year:04d}-{month:02d}-{last_day:02d}"
    except Exception:
        return period


def _derived_total_hours(period: str, utilisation_percent: float) -> float:
    return round(_capacity_hours_for_period(period) * utilisation_percent / 100.0, 1)


def _derived_billable_hours(period: str, utilisation_percent: float, classification: str) -> float:
    if classification == "bench":
        return 0.0
    return _derived_total_hours(period, utilisation_percent)


def _meaningful_tokens(name: str) -> set[str]:
    """Extract meaningful tokens (words ≥3 chars) from a name, stripping onsite/offshore."""
    cleaned = name.lower().strip()
    cleaned = re.sub(r"\s*[-(]?\s*(onsite|offshore)\s*[)]?\s*$", "", cleaned)
    return {w for w in re.split(r"[^a-z]+", cleaned) if len(w) >= 3}


def _token_signature(name: str) -> str:
    return " ".join(sorted(_meaningful_tokens(name)))


def _build_employee_maps(
    employees: list[Employee],
) -> tuple[dict[str, Employee], dict[str, Employee], dict[str, list[Employee]], dict[str, Employee]]:
    exact: dict[str, Employee] = {}
    normalised: dict[str, Employee] = {}
    token_index: dict[str, list[Employee]] = {}
    token_signature_map: dict[str, Employee] = {}
    for employee in employees:
        exact.setdefault(employee.name.strip().lower(), employee)
        normalised.setdefault(_normalise_name(employee.name), employee)
        signature = _token_signature(employee.name)
        if signature and signature not in token_signature_map:
            token_signature_map[signature] = employee
        for token in _meaningful_tokens(employee.name):
            token_index.setdefault(token, []).append(employee)
    return exact, normalised, token_index, token_signature_map


def _fuzzy_match_employee(
    excel_name: str,
    token_index: dict[str, list[Employee]],
    threshold: float = 0.6,
) -> Optional[Employee]:
    """Score candidates by token overlap; return best match above threshold."""
    query_tokens = _meaningful_tokens(excel_name)
    if not query_tokens:
        return None
    candidates: dict[str, tuple[Employee, int]] = {}
    for token in query_tokens:
        for emp in token_index.get(token, []):
            eid = str(emp.id)
            if eid not in candidates:
                candidates[eid] = (emp, 0)
            candidates[eid] = (emp, candidates[eid][1] + 1)

    best_emp: Optional[Employee] = None
    best_score = 0.0
    for emp, common in candidates.values():
        emp_tokens = _meaningful_tokens(emp.name)
        denom = max(len(query_tokens), len(emp_tokens))
        score = common / denom if denom else 0.0
        if score > best_score:
            best_score = score
            best_emp = emp

    return best_emp if best_score >= threshold else None


def _resolve_employee_override(
    employee_name: str,
    overrides: dict[str, Optional[str]],
    emp_by_id: dict[str, Employee],
) -> tuple[bool, Optional[Employee]]:
    if employee_name not in overrides:
        return False, None
    override_id = overrides[employee_name]
    if override_id is None:
        return True, None
    if override_id in emp_by_id:
        return True, emp_by_id[override_id]
    return False, None


def _unique_token_candidate(
    employee_name: str,
    token_index: dict[str, list[Employee]],
) -> Optional[Employee]:
    query_tokens = _meaningful_tokens(employee_name)
    if not query_tokens:
        return None
    candidate_ids: set[str] = set()
    candidate_map: dict[str, Employee] = {}
    for token in query_tokens:
        for employee in token_index.get(token, []):
            employee_id = str(employee.id)
            candidate_ids.add(employee_id)
            candidate_map[employee_id] = employee
    if len(candidate_ids) == 1:
        candidate = candidate_map[next(iter(candidate_ids))]
        candidate_tokens = _meaningful_tokens(candidate.name)
        if query_tokens.issubset(candidate_tokens):
            return candidate
    return None


def _resolve_employee_from_maps(
    employee_name: str,
    exact_map: dict[str, Employee],
    normalised_map: dict[str, Employee],
    token_index: Optional[dict[str, list[Employee]]] = None,
    token_signature_map: Optional[dict[str, Employee]] = None,
) -> Optional[Employee]:
    exact_key = employee_name.strip().lower()
    normalised_key = _normalise_name(employee_name)
    result = exact_map.get(exact_key) or normalised_map.get(normalised_key)
    if result is None and token_signature_map is not None:
        result = token_signature_map.get(_token_signature(employee_name))
    if result is None and token_index is not None:
        result = _fuzzy_match_employee(employee_name, token_index, threshold=0.6)
    if result is None and token_index is not None:
        result = _unique_token_candidate(employee_name, token_index)
    return result


def _load_excel_match_overrides() -> dict[str, Optional[str]]:
    overrides_path = Path(__file__).resolve().parent.parent.parent / "seed" / "match_overrides.json"
    if not overrides_path.exists():
        return {}
    import json as _json
    return _json.loads(overrides_path.read_text())


def _explicit_unmatched_override_names(overrides: dict[str, Optional[str]]) -> set[str]:
    return {name for name, employee_id in overrides.items() if employee_id is None}


async def _build_employee_resolver(branch_location_id: str) -> Callable[[str], Optional[Employee]]:
    branch_employees = await Employee.find(
        Employee.location_id == branch_location_id,
        Employee.is_active == True,
    ).to_list()
    all_active_employees = await Employee.find(
        Employee.is_active == True,
    ).to_list()
    branch_exact, branch_normalised, branch_token_index, branch_token_signature_map = _build_employee_maps(branch_employees)
    global_exact, global_normalised, global_token_index, global_token_signature_map = _build_employee_maps(all_active_employees)
    emp_by_id = {str(e.id): e for e in all_active_employees}

    # Load manual overrides from seed/match_overrides.json if present
    overrides = _load_excel_match_overrides()

    def resolve(employee_name: str) -> Optional[Employee]:
        # Manual override takes priority
        override_hit, override_employee = _resolve_employee_override(employee_name, overrides, emp_by_id)
        if override_hit:
            return override_employee
        # Ignore stale override IDs and continue with live HRMS matching.
        branch_exact_match = branch_exact.get(employee_name.strip().lower()) or branch_normalised.get(_normalise_name(employee_name))
        if branch_exact_match is not None:
            return branch_exact_match
        branch_signature_match = branch_token_signature_map.get(_token_signature(employee_name))
        if branch_signature_match is not None:
            return branch_signature_match
        return _resolve_employee_from_maps(
            employee_name,
            global_exact,
            global_normalised,
            global_token_index,
            global_token_signature_map,
        )

    return resolve


def _effective_excel_employee_id(
    row: ExcelUtilisationReport,
    resolved_employee: Optional[Employee],
    explicit_unmatched_names: Optional[set[str]] = None,
) -> Optional[str]:
    if explicit_unmatched_names and row.employee_name in explicit_unmatched_names:
        return None
    stored_id = row.employee_id if row.employee_id and not row.employee_id.startswith("unmatched:") else None
    return stored_id or (str(resolved_employee.id) if resolved_employee else None)


async def _enrich_excel_people(
    rows: list[ExcelUtilisationReport],
    branch_location_id: str,
) -> list[dict]:
    resolve_employee = await _build_employee_resolver(branch_location_id)
    explicit_unmatched_names = _explicit_unmatched_override_names(_load_excel_match_overrides())

    resolved_ids: list[str] = []
    enriched_people: list[dict] = []
    for row in rows:
        resolved_employee = resolve_employee(row.employee_name)
        resolved_employee_id = _effective_excel_employee_id(row, resolved_employee, explicit_unmatched_names)
        if resolved_employee_id:
            resolved_ids.append(resolved_employee_id)
        enriched_people.append(
            {
                "row": row,
                "resolved_employee_id": resolved_employee_id,
            }
        )

    employee_map: dict[str, Employee] = {}
    if resolved_ids:
        employees = await Employee.find(
            {"_id": {"$in": [ObjectId(employee_id) for employee_id in resolved_ids if ObjectId.is_valid(employee_id)]}}
        ).to_list()
        employee_map = {str(employee.id): employee for employee in employees}

    dept_ids = {employee.department_id for employee in employee_map.values() if employee.department_id}
    departments = (
        await Department.find(
            {"_id": {"$in": [ObjectId(dept_id) for dept_id in dept_ids if ObjectId.is_valid(dept_id)]}}
        ).to_list()
        if dept_ids
        else []
    )
    dept_map = {str(department.id): department.name for department in departments}

    location_ids = {employee.location_id for employee in employee_map.values() if employee.location_id}
    locations = (
        await Location.find(
            {"_id": {"$in": [ObjectId(location_id) for location_id in location_ids if ObjectId.is_valid(location_id)]}}
        ).to_list()
        if location_ids
        else []
    )
    location_map = {str(location.id): location for location in locations}

    for person in enriched_people:
        resolved_employee_id = person["resolved_employee_id"]
        resolved_employee = employee_map.get(resolved_employee_id) if resolved_employee_id else None
        department_name = dept_map.get(resolved_employee.department_id, "Unknown") if resolved_employee else "Unknown"
        location = location_map.get(resolved_employee.location_id) if resolved_employee else None
        person["resolved_employee"] = resolved_employee
        person["department_name"] = department_name
        person["location"] = location

    return enriched_people


def _build_excel_timesheet_entry(
    row: ExcelUtilisationReport,
    resolved_employee_id: Optional[str],
) -> dict:
    employee_ref = resolved_employee_id or _excel_employee_ref(row.employee_name)
    total_hours = _derived_total_hours(row.period, row.utilisation_percent)
    billable_hours = _derived_billable_hours(row.period, row.utilisation_percent, row.classification)
    return {
        "id": f"excel:{row.period}:{_normalise_employee_key(row.employee_name, resolved_employee_id)}",
        "employee_id": employee_ref,
        "employee_name": row.employee_name,
        "project_id": EXCEL_TIMESHEET_PROJECT_ID,
        "project_name": EXCEL_TIMESHEET_PROJECT_NAME,
        "date": _period_date(row.period),
        "hours": total_hours,
        "is_billable": billable_hours > 0,
        "description": f"Imported monthly utilisation snapshot ({row.utilisation_percent:.1f}% utilised, {row.availability_percent:.1f}% available)",
        "status": "approved",
        "source": "excel_upload",
        "period": row.period,
        "created_at": row.uploaded_at.isoformat(),
        "updated_at": row.uploaded_at.isoformat(),
    }


async def parse_and_store_excel(
    file_content: bytes,
    filename: str,
    branch_location_id: str,
    user_id: str,
) -> dict:
    workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    sheet = _select_sheet(workbook)
    resolve_employee = await _build_employee_resolver(branch_location_id)

    all_rows = list(sheet.iter_rows(values_only=True))
    month_col_map, header_row_idx = _parse_ytpl_sheet(sheet)

    if not month_col_map:
        return {"error": "No month columns found. Expected datetime headers in the sheet.", "total_rows": 0}

    header = all_rows[header_row_idx]

    name_col = None
    practice_col = None
    function_col = None
    status_col = None

    for col_index, cell in enumerate(header):
        if col_index in month_col_map:
            continue
        value = _normalise_text(cell).lower()
        if value == "name":
            name_col = col_index
        elif value == "practice":
            practice_col = col_index
        elif value == "function":
            function_col = col_index
        elif value == "status":
            status_col = col_index

    if name_col is None:
        name_col = 2
    if practice_col is None:
        practice_col = 6
    if function_col is None:
        function_col = 5
    if status_col is None:
        status_col = 3

    # Detect "Revised Availability" and "Remarks" columns for fallback parsing
    revised_avail_col: Optional[int] = None
    remarks_col: Optional[int] = None
    for col_index, cell in enumerate(header):
        if col_index in month_col_map:
            continue
        value = _normalise_text(cell).lower()
        if "revised" in value and "avail" in value:
            revised_avail_col = col_index
        elif value == "remarks":
            remarks_col = col_index

    # Determine the "current" period for fallback rows (latest month in the sheet)
    fallback_period = max(month_col_map.values()) if month_col_map else "2026-03"

    batch_id = secrets.token_hex(8)
    now = datetime.now(timezone.utc)
    periods_found: set[str] = set()
    docs: list[ExcelUtilisationReport] = []
    total_rows = 0
    matched_rows = 0

    for row in all_rows[header_row_idx + 1 :]:
        if not row or len(row) <= name_col:
            continue

        employee_name = _normalise_text(row[name_col])
        if not employee_name or employee_name.lower() in {"name", "total", "grand total"}:
            continue

        if status_col is not None and len(row) > status_col:
            status_value = _normalise_text(row[status_col]).lower()
            if status_value in {"left", "inactive", "resigned", "terminated"}:
                continue

        practice = (
            _normalise_text(row[practice_col])
            if practice_col is not None and len(row) > practice_col and row[practice_col]
            else None
        )
        function_name = (
            _normalise_text(row[function_col])
            if function_col is not None and len(row) > function_col and row[function_col]
            else None
        )
        department = practice or function_name

        resolved_employee = resolve_employee(employee_name)
        if resolved_employee and resolved_employee.location_id != branch_location_id:
            continue
        employee_id = str(resolved_employee.id) if resolved_employee else f"unmatched:{_normalise_name(employee_name)}"
        total_rows += 1
        if resolved_employee:
            matched_rows += 1

        # --- Try month columns first (fractions per month) ---
        row_had_month_data = False
        for col_index, period in month_col_map.items():
            if col_index >= len(row):
                continue
            cell_value = row[col_index]
            if cell_value is None:
                continue

            try:
                availability_fraction = float(cell_value)
            except (TypeError, ValueError):
                continue

            availability_percent = round(max(0.0, min(100.0, availability_fraction * 100)), 2)
            utilisation_percent = round(100.0 - availability_percent, 2)
            periods_found.add(period)
            row_had_month_data = True

            docs.append(
                ExcelUtilisationReport(
                    branch_location_id=branch_location_id,
                    upload_batch_id=batch_id,
                    uploaded_at=now,
                    uploaded_by=user_id,
                    filename=filename,
                    employee_name=employee_name,
                    employee_email=resolved_employee.email if resolved_employee else None,
                    employee_id=employee_id,
                    department=department,
                    period=period,
                    availability_percent=availability_percent,
                    utilisation_percent=utilisation_percent,
                    classification=_classify(availability_percent),
                )
            )

        if row_had_month_data:
            continue

        # --- Fallback: use "Revised Availability" col or parse "Remarks" text ---
        availability_fraction: Optional[float] = None

        if revised_avail_col is not None and revised_avail_col < len(row) and row[revised_avail_col] is not None:
            try:
                availability_fraction = float(row[revised_avail_col])
            except (TypeError, ValueError):
                pass

        if availability_fraction is None and remarks_col is not None and remarks_col < len(row):
            availability_fraction = _parse_remarks_availability(row[remarks_col])

        if availability_fraction is None:
            continue  # No usable availability data for this employee

        availability_percent = round(max(0.0, min(100.0, availability_fraction * 100)), 2)
        utilisation_percent = round(100.0 - availability_percent, 2)
        periods_found.add(fallback_period)

        docs.append(
            ExcelUtilisationReport(
                branch_location_id=branch_location_id,
                upload_batch_id=batch_id,
                uploaded_at=now,
                uploaded_by=user_id,
                filename=filename,
                employee_name=employee_name,
                employee_email=resolved_employee.email if resolved_employee else None,
                employee_id=employee_id,
                department=department,
                period=fallback_period,
                availability_percent=availability_percent,
                utilisation_percent=utilisation_percent,
                classification=_classify(availability_percent),
            )
        )

    periods = sorted(periods_found)
    if docs and periods:
        await ExcelUtilisationReport.find(
            ExcelUtilisationReport.branch_location_id == branch_location_id,
            {"period": {"$in": periods}},
        ).delete()
        await ExcelUtilisationReport.insert_many(docs)

    upload_log = ExcelUploadLog(
        batch_id=batch_id,
        branch_location_id=branch_location_id,
        uploaded_by=user_id,
        filename=filename,
        total_rows=total_rows,
        matched_rows=matched_rows,
        periods=periods,
        uploaded_at=now,
    )
    await upload_log.insert()

    return {
        "batch_id": batch_id,
        "total_rows": total_rows,
        "matched_rows": matched_rows,
        "periods": periods,
        "rows_stored": len(docs),
    }


async def _get_intercompany_rows_as_utilisation(
    period: str,
    branch_location_id: str,
) -> list["ExcelUtilisationReport"]:
    """
    For periods with no YTPL availability upload (e.g. 2026-03),
    synthesise virtual ExcelUtilisationReport rows from inter-company
    ProjectAllocation records so the Excel toggle still works.
    """
    allocations = await ProjectAllocation.find(
        ProjectAllocation.period == period,
        ProjectAllocation.source_system == "intercompany_excel",
        ProjectAllocation.is_deleted != True,
    ).to_list()

    if not allocations:
        return []

    # Aggregate per employee (sum allocation %)
    from collections import defaultdict as _dd
    emp_alloc: dict[str, float] = _dd(float)
    emp_name: dict[str, str] = {}
    for alloc in allocations:
        emp_alloc[alloc.employee_id] += alloc.allocation_percentage
        emp_name[alloc.employee_id] = alloc.employee_name

    # Load employees to get branch filter
    emp_ids = list(emp_alloc.keys())
    employees = await Employee.find(
        {"_id": {"$in": [ObjectId(eid) for eid in emp_ids if ObjectId.is_valid(eid)]}},
        Employee.location_id == branch_location_id,
        Employee.is_active == True,
    ).to_list()
    branch_emp_ids = {str(e.id) for e in employees}

    now = datetime.now(timezone.utc)
    virtual_rows = []
    for emp_id, total_alloc_pct in emp_alloc.items():
        if emp_id not in branch_emp_ids:
            continue
        utilisation_percent = round(min(100.0, total_alloc_pct), 2)
        availability_percent = round(max(0.0, 100.0 - utilisation_percent), 2)
        virtual_rows.append(
            ExcelUtilisationReport(
                branch_location_id=branch_location_id,
                upload_batch_id="intercompany_synthetic",
                uploaded_at=now,
                uploaded_by="system",
                filename="Inter-company sheet",
                employee_name=emp_name.get(emp_id, emp_id),
                employee_id=emp_id,
                period=period,
                availability_percent=availability_percent,
                utilisation_percent=utilisation_percent,
                classification=_classify(availability_percent),
            )
        )
    return virtual_rows


async def _get_combined_excel_rows_for_period(
    period: str,
    branch_location_id: str,
) -> list["ExcelUtilisationReport"]:
    rows = await ExcelUtilisationReport.find(
        ExcelUtilisationReport.branch_location_id == branch_location_id,
        ExcelUtilisationReport.period == period,
    ).to_list()
    rows = _dedupe_latest_rows(rows)

    intercompany_rows = await _get_intercompany_rows_as_utilisation(period, branch_location_id)
    if not intercompany_rows:
        return rows

    existing_keys = {_normalise_employee_key(row.employee_name, row.employee_id) for row in rows}
    for row in intercompany_rows:
        key = _normalise_employee_key(row.employee_name, row.employee_id)
        if key not in existing_keys:
            rows.append(row)
            existing_keys.add(key)
    return rows


async def _get_all_excel_employee_rows(branch_location_id: str) -> list["ExcelUtilisationReport"]:
    rows = await ExcelUtilisationReport.find(
        ExcelUtilisationReport.branch_location_id == branch_location_id,
    ).to_list()
    rows = _dedupe_latest_rows(rows)

    intercompany_periods = sorted(
        {
            allocation.period
            for allocation in await ProjectAllocation.find(
                ProjectAllocation.source_system == "intercompany_excel",
                ProjectAllocation.is_deleted != True,
            ).to_list()
        }
    )
    existing_keys = {_normalise_employee_key(row.employee_name, row.employee_id) for row in rows}
    for period in intercompany_periods:
        for row in await _get_intercompany_rows_as_utilisation(period, branch_location_id):
            key = _normalise_employee_key(row.employee_name, row.employee_id)
            if key not in existing_keys:
                rows.append(row)
                existing_keys.add(key)

    return rows


def _summarise_excel_people(enriched_people: list[dict]) -> tuple[int, int, int]:
    unique_people: dict[str, tuple[Optional[Employee], ExcelUtilisationReport]] = {}
    for person in enriched_people:
        resolved_employee_id = person.get("resolved_employee_id")
        resolved_employee = person.get("resolved_employee")
        row = person.get("row")
        if row is None:
            continue
        key = resolved_employee_id or _excel_employee_ref(row.employee_name)
        unique_people.setdefault(key, (resolved_employee, row))

    total = len(unique_people)
    active_count = sum(
        1
        for resolved_employee, _row in unique_people.values()
        if getattr(resolved_employee, "is_active", True)
    )
    inactive_count = total - active_count
    return total, active_count, inactive_count


async def get_excel_dashboard(
    period: str,
    branch_location_id: str,
) -> Optional[dict]:
    rows = await _get_combined_excel_rows_for_period(period, branch_location_id)

    if not rows:
        return None

    total = len(rows)
    fully_billed = sum(1 for row in rows if row.classification == "fully_billed")
    partially_billed = sum(1 for row in rows if row.classification == "partially_billed")
    bench_count = sum(1 for row in rows if row.classification == "bench")
    billable_count = fully_billed + partially_billed

    avg_util = round(sum(row.utilisation_percent for row in rows) / total, 2) if total else 0.0
    avg_avail = round(sum(row.availability_percent for row in rows) / total, 2) if total else 0.0

    classification_breakdown = [
        {
            "classification": "fully_billed",
            "count": fully_billed,
            "percent": round(fully_billed / total * 100, 2) if total else 0,
        },
        {
            "classification": "partially_billed",
            "count": partially_billed,
            "percent": round(partially_billed / total * 100, 2) if total else 0,
        },
        {
            "classification": "bench",
            "count": bench_count,
            "percent": round(bench_count / total * 100, 2) if total else 0,
        },
    ]

    resource_availability = {
        "available": bench_count,
        "fully_allocated": billable_count,
        "over_allocated": 0,
    }

    from app.services.dashboard_service import _previous_periods

    trend_periods = _previous_periods(period, 6)
    trend_rows = await ExcelUtilisationReport.find(
        ExcelUtilisationReport.branch_location_id == branch_location_id,
        {"period": {"$in": trend_periods}},
    ).to_list()

    grouped_rows: dict[str, list[ExcelUtilisationReport]] = defaultdict(list)
    for trend_period in trend_periods:
        grouped_rows[trend_period] = _dedupe_latest_rows(
            [row for row in trend_rows if row.period == trend_period]
        )

    trend = []
    for trend_period in trend_periods:
        group = grouped_rows.get(trend_period, [])
        count = len(group)
        avg_period_util = round(sum(row.utilisation_percent for row in group) / count, 2) if count else 0.0
        trend.append(
            {
                "period": trend_period,
                "headcount": count,
                "utilisation_percent": avg_period_util,
                "billable_percent": avg_period_util,
            }
        )

    # Top consuming projects from inter-company allocations for this period
    ic_allocations = await ProjectAllocation.find(
        ProjectAllocation.period == period,
        ProjectAllocation.source_system == "intercompany_excel",
        ProjectAllocation.is_deleted != True,
    ).to_list()
    project_headcount: dict[str, dict] = {}
    for alloc in ic_allocations:
        key = alloc.project_id
        if key not in project_headcount:
            project_headcount[key] = {
                "project_name": alloc.project_name,
                "client_name": alloc.client_name,
                "employee_count": 0,
                "total_hours": 0.0,
            }
        project_headcount[key]["employee_count"] += 1
        project_headcount[key]["total_hours"] += alloc.allocated_days * 8.0
    top_consuming_projects = sorted(
        [{"project_name": v["project_name"], "client_name": v["client_name"],
          "employee_count": v["employee_count"], "total_hours": round(v["total_hours"], 1)}
         for v in project_headcount.values()],
        key=lambda x: x["employee_count"],
        reverse=True,
    )[:10]

    return {
        "period": period,
        "data_source": "excel",
        "total_active_employees": total,
        "billable_count": billable_count,
        "non_billable_count": 0,
        "bench_count": bench_count,
        "overall_utilisation_percent": avg_util,
        "overall_billable_percent": round(100 - avg_avail, 2),
        "top_consuming_projects": top_consuming_projects,
        "resource_availability": resource_availability,
        "classification_breakdown": classification_breakdown,
        "trend": trend,
    }


async def get_excel_resource_rows(
    period: str,
    branch_location_id: str,
    search: Optional[str] = None,
    classification: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    rows = await _get_combined_excel_rows_for_period(period, branch_location_id)

    if not rows:
        return {"period": period, "data_source": "excel", "entries": [], "total": 0}

    enriched_people = await _enrich_excel_people(rows, branch_location_id)
    for person in enriched_people:
        row = person["row"]
        resolved_employee = person.get("resolved_employee")
        person["display_name"] = resolved_employee.name if resolved_employee else row.employee_name
        person["derived_available_days"] = round(_working_days_in_period(period) * row.availability_percent / 100, 1)
        person["derived_total_hours"] = _derived_total_hours(period, row.utilisation_percent)
        person["derived_billable_hours"] = _derived_billable_hours(period, row.utilisation_percent, row.classification)

    if classification:
        enriched_people = [person for person in enriched_people if person["row"].classification == classification]

    resolved_employee_ids = [person["resolved_employee_id"] for person in enriched_people if person["resolved_employee_id"]]

    allocations = (
        await ProjectAllocation.find(
            ProjectAllocation.period == period,
            {"employee_id": {"$in": resolved_employee_ids}},
            ProjectAllocation.source_system == "intercompany_excel",
            ProjectAllocation.is_deleted != True,
        ).to_list()
        if resolved_employee_ids
        else []
    )
    allocations_by_employee: dict[str, list[ProjectAllocation]] = defaultdict(list)
    for allocation in allocations:
        allocations_by_employee[allocation.employee_id].append(allocation)

    reporting_relationships = (
        await ReportingRelationship.find(
            {"employee_id": {"$in": resolved_employee_ids}},
            ReportingRelationship.type == "PRIMARY",
            ReportingRelationship.is_deleted != True,
        ).to_list()
        if resolved_employee_ids
        else []
    )
    manager_ids = [relationship.manager_id for relationship in reporting_relationships]
    managers = (
        await Employee.find(
            {"_id": {"$in": [ObjectId(manager_id) for manager_id in manager_ids if ObjectId.is_valid(manager_id)]}}
        ).to_list()
        if manager_ids
        else []
    )
    manager_name_by_id = {str(manager.id): manager.name for manager in managers}
    line_manager_by_employee: dict[str, str] = {}
    for relationship in reporting_relationships:
        line_manager_by_employee[relationship.employee_id] = manager_name_by_id.get(relationship.manager_id, "No Manager")

    effective_timesheet_entries = await get_effective_timesheet_entries(
        period=period,
        employee_ids=resolved_employee_ids if resolved_employee_ids else [],
    )
    employee_project_hours = summarise_hours_by_employee_project(effective_timesheet_entries)
    employee_total_hours = summarise_hours_by_employee(effective_timesheet_entries)

    merged_rows: list[dict] = []
    for person in enriched_people:
        row = person["row"]
        resolved_employee_id = person["resolved_employee_id"]
        employee_ref = resolved_employee_id or _excel_employee_ref(row.employee_name)
        line_manager = line_manager_by_employee.get(resolved_employee_id, "No Manager") if resolved_employee_id else "No Manager"
        row_allocations = allocations_by_employee.get(resolved_employee_id, []) if resolved_employee_id else []

        if row_allocations:
            for allocation in row_allocations:
                project_hours = employee_project_hours.get(resolved_employee_id, {}).get(
                    allocation.project_id,
                    {"billable": 0.0, "non_billable": 0.0},
                )
                merged_rows.append(
                    {
                        "employee_id": employee_ref,
                        "employee_name": person["display_name"],
                        "line_manager": line_manager,
                        "project_name": allocation.project_name,
                        "client_name": allocation.client_name,
                        "department": row.department,
                        "allocation_percentage": allocation.allocation_percentage,
                        "availability_percent": row.availability_percent,
                        "utilisation_percent": row.utilisation_percent,
                        "billable_hours": round(project_hours["billable"], 1),
                        "non_billable_hours": round(project_hours["non_billable"], 1),
                        "classification": row.classification,
                        "available_days": allocation.available_days,
                    }
                )
        else:
            employee_hours = employee_total_hours.get(
                resolved_employee_id,
                {"billable": person["derived_billable_hours"], "non_billable": round(person["derived_total_hours"] - person["derived_billable_hours"], 1)},
            )
            merged_rows.append(
                {
                    "employee_id": employee_ref,
                    "employee_name": person["display_name"],
                    "line_manager": line_manager,
                    "project_name": None,
                    "client_name": None,
                    "department": row.department,
                    "allocation_percentage": round(row.utilisation_percent, 1),
                    "availability_percent": row.availability_percent,
                    "utilisation_percent": row.utilisation_percent,
                    "billable_hours": round(employee_hours["billable"], 1),
                    "non_billable_hours": round(employee_hours["non_billable"], 1),
                    "classification": row.classification,
                    "available_days": person["derived_available_days"],
                }
            )

    if search:
        search_lower = search.lower()
        merged_rows = [
            merged_row
            for merged_row in merged_rows
            if search_lower in merged_row["employee_name"].lower()
            or search_lower in (merged_row["line_manager"] or "").lower()
            or search_lower in (merged_row["project_name"] or "").lower()
            or search_lower in (merged_row["client_name"] or "").lower()
            or search_lower in (merged_row["department"] or "").lower()
        ]

    total = len(merged_rows)
    start = (page - 1) * page_size
    page_rows = merged_rows[start : start + page_size]

    return {"period": period, "data_source": "excel", "entries": page_rows, "total": total}


async def get_excel_timesheets(
    period: str,
    branch_location_id: str,
    employee_id: Optional[str] = None,
    project_id: Optional[str] = None,
    status: Optional[str] = None,
    is_billable: Optional[bool] = None,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    rows = await ExcelUtilisationReport.find(
        ExcelUtilisationReport.branch_location_id == branch_location_id,
        ExcelUtilisationReport.period == period,
    ).to_list()

    deduped_rows = _dedupe_latest_rows(rows)

    resolve_employee = await _build_employee_resolver(branch_location_id)
    explicit_unmatched_names = _explicit_unmatched_override_names(_load_excel_match_overrides())
    all_entries = []
    for row in deduped_rows:
        resolved_employee = resolve_employee(row.employee_name)
        resolved_employee_id = _effective_excel_employee_id(row, resolved_employee, explicit_unmatched_names)
        all_entries.append(_build_excel_timesheet_entry(row, resolved_employee_id))

    if employee_id:
        all_entries = [entry for entry in all_entries if entry["employee_id"] == employee_id]
    if project_id:
        all_entries = [entry for entry in all_entries if entry["project_id"] == project_id]
    if status:
        all_entries = [entry for entry in all_entries if entry["status"] == status]
    if is_billable is not None:
        all_entries = [entry for entry in all_entries if entry["is_billable"] == is_billable]

    all_entries.sort(key=lambda entry: (entry["date"], entry["employee_name"]), reverse=True)
    total = len(all_entries)
    start = (page - 1) * page_size
    page_entries = all_entries[start : start + page_size]

    total_hours = round(sum(entry["hours"] for entry in deduped_rows_to_entries(deduped_rows, resolve_employee, period)), 1)
    billable_hours = round(sum(entry["billable_hours"] for entry in deduped_rows_to_entries(deduped_rows, resolve_employee, period)), 1)
    billable_entries = [entry for entry in deduped_rows_to_entries(deduped_rows, resolve_employee, period) if entry["is_billable"]]
    non_billable_entries = [entry for entry in deduped_rows_to_entries(deduped_rows, resolve_employee, period) if not entry["is_billable"]]

    employee_options = sorted(
        [{"id": entry["employee_id"], "name": entry["employee_name"]} for entry in deduped_rows_to_entries(deduped_rows, resolve_employee, period)],
        key=lambda item: item["name"].lower(),
    )

    latest_period = await _get_latest_excel_period(branch_location_id)

    return {
        "entries": page_entries,
        "total": total,
        "period": period,
        "is_locked": False,
        "latest_period": latest_period if total == 0 else None,
        "summary": {
            "total_hours": total_hours,
            "billable_hours": billable_hours,
            "billable_percent": round((billable_hours / total_hours * 100) if total_hours > 0 else 0, 1),
            "employee_count": len({entry["employee_id"] for entry in deduped_rows_to_entries(deduped_rows, resolve_employee, period)}),
            "project_count": 1 if deduped_rows else 0,
            "billable_employee_count": len({entry["employee_id"] for entry in billable_entries}),
            "non_billable_employee_count": len({entry["employee_id"] for entry in non_billable_entries}),
        },
        "filter_options": {
            "employees": employee_options,
            "projects": [{"id": EXCEL_TIMESHEET_PROJECT_ID, "name": EXCEL_TIMESHEET_PROJECT_NAME}] if deduped_rows else [],
        },
        "data_source": "excel",
    }


def deduped_rows_to_entries(
    rows: list[ExcelUtilisationReport],
    resolve_employee: Callable[[str], Optional[Employee]],
    period: str,
) -> list[dict]:
    explicit_unmatched_names = _explicit_unmatched_override_names(_load_excel_match_overrides())
    entries = []
    for row in rows:
        resolved_employee = resolve_employee(row.employee_name)
        resolved_employee_id = _effective_excel_employee_id(row, resolved_employee, explicit_unmatched_names)
        entry = _build_excel_timesheet_entry(row, resolved_employee_id)
        entry["billable_hours"] = _derived_billable_hours(period, row.utilisation_percent, row.classification)
        entries.append(entry)
    return entries


async def get_excel_analytics(
    branch_location_id: str,
    period: Optional[str] = None,
) -> Optional[dict]:
    target_period = period or await _get_latest_excel_period(branch_location_id)
    if not target_period:
        return None

    rows = await _get_combined_excel_rows_for_period(target_period, branch_location_id)
    if not rows:
        return None

    employee_ids = [row.employee_id for row in rows if row.employee_id]
    employees = (
        await Employee.find(
            {"_id": {"$in": [ObjectId(employee_id) for employee_id in employee_ids if ObjectId.is_valid(employee_id)]}},
            Employee.is_active == True,
        ).to_list()
        if employee_ids
        else []
    )
    employee_map = {str(employee.id): employee for employee in employees}
    employee_ids = [employee_id for employee_id in employee_ids if employee_id in employee_map]

    all_employees = await Employee.find(Employee.is_active == True).to_list()
    all_emp_map = {str(employee.id): employee for employee in all_employees}

    departments = await Department.find_all().to_list()
    dept_map = {str(department.id): department for department in departments}
    locations = await Location.find_all().to_list()
    loc_map = {str(location.id): location for location in locations}

    level_order = ["intern", "junior", "mid", "senior", "lead", "manager", "head", "director", "vp", "c-suite"]

    level_counts = defaultdict(int)
    department_ids = set()
    for employee_id in employee_ids:
        employee = employee_map.get(employee_id)
        if not employee:
            continue
        level_counts[employee.level] += 1
        if employee.department_id:
            department_ids.add(employee.department_id)

    level_breakdown = [
        {"level": level, "count": level_counts.get(level, 0)}
        for level in level_order
        if level_counts.get(level, 0) > 0
    ]

    reporting_relationships = (
        await ReportingRelationship.find(
            {"employee_id": {"$in": employee_ids}},
            ReportingRelationship.is_deleted != True,
        ).to_list()
        if employee_ids
        else []
    )
    primary_rels = [relationship for relationship in reporting_relationships if relationship.type == "PRIMARY"]
    manager_reports: dict[str, list[str]] = defaultdict(list)
    parent_of: dict[str, str] = {}
    for relationship in primary_rels:
        manager_reports[relationship.manager_id].append(relationship.employee_id)
        parent_of[relationship.employee_id] = relationship.manager_id

    span_data = []
    for manager_id, reports in manager_reports.items():
        if manager_id not in employee_ids:
            continue
        manager = all_emp_map.get(manager_id)
        if not manager:
            continue
        count = len(reports)
        span_data.append(
            {
                "manager_id": manager_id,
                "manager_name": manager.name,
                "designation": manager.designation,
                "direct_report_count": count,
                "is_outlier": count < 2 or count > 10,
            }
        )

    def get_depth(employee_id: str, visited: Optional[set[str]] = None) -> int:
        if visited is None:
            visited = set()
        if employee_id in visited:
            return 0
        visited.add(employee_id)
        parent = parent_of.get(employee_id)
        if not parent or parent not in employee_ids:
            return 0
        return 1 + get_depth(parent, visited)

    hierarchy_depth = 0
    for employee_id in employee_ids:
        hierarchy_depth = max(hierarchy_depth, get_depth(employee_id))

    dept_has_manager = set()
    for employee_id in employee_ids:
        employee = employee_map.get(employee_id)
        if not employee:
            continue
        if employee.level in ("manager", "head", "director", "vp", "c-suite") and employee.department_id:
            dept_has_manager.add(employee.department_id)

    departments_without_manager = sorted(
        dept_map[department_id].name
        for department_id in (department_ids - dept_has_manager)
        if department_id in dept_map
    )

    cross_reports = []
    for relationship in reporting_relationships:
        if relationship.employee_id not in employee_ids:
            continue
        employee = all_emp_map.get(relationship.employee_id)
        manager = all_emp_map.get(relationship.manager_id)
        if not employee or not manager or manager.location_id == branch_location_id:
            continue
        manager_location = loc_map.get(manager.location_id)
        cross_reports.append(
            {
                "employee_id": relationship.employee_id,
                "employee_name": employee.name,
                "employee_designation": employee.designation,
                "external_manager_id": relationship.manager_id,
                "external_manager_name": manager.name,
                "external_manager_location": manager_location.city if manager_location else "Unknown",
                "relationship_type": relationship.type,
            }
        )

    allocations = (
        await ProjectAllocation.find(
            ProjectAllocation.period == target_period,
            {"employee_id": {"$in": employee_ids}},
            ProjectAllocation.source_system == "intercompany_excel",
            ProjectAllocation.is_deleted != True,
        ).to_list()
        if employee_ids
        else []
    )

    client_counts: dict[str, set[str]] = defaultdict(set)
    project_member_count: dict[str, set[str]] = defaultdict(set)
    project_snapshot: dict[str, dict] = {}
    for allocation in allocations:
        client_name = allocation.client_name or "General"
        client_counts[client_name].add(allocation.employee_id)
        project_member_count[allocation.project_id].add(allocation.employee_id)
        project_snapshot.setdefault(
            allocation.project_id,
            {
                "id": allocation.project_id,
                "name": allocation.project_name,
                "status": "ACTIVE",
                "member_count": 0,
                "client_name": client_name,
            },
        )

    client_breakdown = sorted(
        [{"client": client_name, "count": len(employee_set)} for client_name, employee_set in client_counts.items()],
        key=lambda item: item["count"],
        reverse=True,
    )

    project_summaries = sorted(
        [
            {
                **summary,
                "member_count": len(project_member_count[project_id]),
            }
            for project_id, summary in project_snapshot.items()
        ],
        key=lambda item: item["member_count"],
        reverse=True,
    )

    trend_periods = sorted(
        {
            row.period
            for row in await ExcelUtilisationReport.find(
                ExcelUtilisationReport.branch_location_id == branch_location_id,
            ).to_list()
        }
    )[-12:]
    monthly_trend = []
    previous_count = 0
    for trend_period in trend_periods:
        trend_rows = await ExcelUtilisationReport.find(
            ExcelUtilisationReport.branch_location_id == branch_location_id,
            ExcelUtilisationReport.period == trend_period,
        ).to_list()
        count = len(_dedupe_latest_rows(trend_rows))
        monthly_trend.append(
            {
                "month": trend_period,
                "new_hires": max(count - previous_count, 0),
                "cumulative": count,
            }
        )
        previous_count = count

    return {
        "period": target_period,
        "data_source": "excel",
        "total_headcount": len(rows),
        "active_count": len(rows),
        "client_breakdown": client_breakdown,
        "level_breakdown": level_breakdown,
        "monthly_trend": monthly_trend,
        "span_of_control": span_data,
        "hierarchy_depth": hierarchy_depth,
        "departments_without_manager": departments_without_manager,
        "cross_reports": cross_reports,
        "projects": project_summaries,
        "orphaned_projects": [],
    }


async def get_excel_bench_pool(
    branch_location_id: str,
    period: Optional[str] = None,
    skill_filter: Optional[str] = None,
    location_filter: Optional[str] = None,
    classification_filter: Optional[str] = None,
    designation_filter: Optional[str] = None,
    utilisation_min: Optional[float] = None,
    utilisation_max: Optional[float] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    target_period = period or await _get_latest_excel_period(branch_location_id)
    if not target_period:
        return {
            "period": None,
            "data_source": "excel",
            "employees": [],
            "total": 0,
            "bench_count": 0,
            "partial_count": 0,
            "avg_bench_days": None,
        }

    rows = await _get_combined_excel_rows_for_period(target_period, branch_location_id)
    rows = [row for row in rows if row.classification in {"bench", "partially_billed"}]

    if classification_filter in {"bench", "partially_billed"}:
        rows = [row for row in rows if row.classification == classification_filter]

    if not rows:
        return {
            "period": target_period,
            "data_source": "excel",
            "employees": [],
            "total": 0,
            "bench_count": 0,
            "partial_count": 0,
            "avg_bench_days": None,
        }

    enriched_people = await _enrich_excel_people(rows, branch_location_id)

    def bench_employee_ref(person: dict) -> str:
        row = person["row"]
        resolved_employee_id = person.get("resolved_employee_id")
        if resolved_employee_id:
            return resolved_employee_id
        if row.employee_id and row.employee_id.startswith("unmatched:"):
            return row.employee_id
        return _excel_employee_ref(row.employee_name)

    matched_employee_ids = sorted(
        {
            person["resolved_employee_id"]
            for person in enriched_people
            if person.get("resolved_employee_id") and person.get("resolved_employee") is not None
        }
    )
    allocation_employee_refs = sorted({bench_employee_ref(person) for person in enriched_people})

    all_skills = (
        await EmployeeSkill.find({"employee_id": {"$in": matched_employee_ids}}).to_list()
        if matched_employee_ids
        else []
    )
    skills_by_emp: dict[str, list[EmployeeSkill]] = defaultdict(list)
    for skill in all_skills:
        skills_by_emp[skill.employee_id].append(skill)

    assignments = (
        await EmployeeProject.find({"employee_id": {"$in": matched_employee_ids}}).to_list()
        if matched_employee_ids
        else []
    )
    project_ids = list({assignment.project_id for assignment in assignments if assignment.project_id})
    projects = (
        await Project.find(
            {"_id": {"$in": [ObjectId(project_id) for project_id in project_ids if ObjectId.is_valid(project_id)]}}
        ).to_list()
        if project_ids
        else []
    )
    project_map = {str(project.id): project for project in projects}

    assignments_by_emp: dict[str, list[dict]] = defaultdict(list)
    employee_project_role: dict[tuple[str, str], str] = {}
    available_from_map: dict[str, Optional[str]] = {}
    for assignment in assignments:
        project = project_map.get(assignment.project_id)
        if not project:
            continue
        end_date_str = project.end_date.strftime("%Y-%m-%d") if project.end_date else None
        assignment_is_deleted = getattr(assignment, "is_deleted", False)
        assignments_by_emp[assignment.employee_id].append(
            {
                "project_id": assignment.project_id,
                "project_name": project.name,
                "status": "COMPLETED" if assignment_is_deleted else project.status,
                "role": assignment.role_in_project,
                "end_date": end_date_str,
                "client_name": project.client_name,
                "is_deleted": assignment_is_deleted,
            }
        )
        if assignment.role_in_project:
            employee_project_role[(assignment.employee_id, assignment.project_id)] = assignment.role_in_project
        if not assignment_is_deleted and project.status == "ACTIVE" and project.end_date:
            end_str = project.end_date.strftime("%Y-%m-%d")
            current = available_from_map.get(assignment.employee_id)
            if current is None or end_str > current:
                available_from_map[assignment.employee_id] = end_str

    allocations = await ProjectAllocation.find(
        {
            "employee_id": {"$in": allocation_employee_refs},
            "allocated_days": {"$gt": 0},
            "source_system": "intercompany_excel",
        }
    ).sort(-ProjectAllocation.period).to_list()

    alloc_last_by_emp: dict[str, list[dict]] = defaultdict(list)
    for allocation in allocations:
        employee_allocations = alloc_last_by_emp[allocation.employee_id]
        if any(existing["project_id"] == allocation.project_id for existing in employee_allocations):
            continue
        employee_allocations.append(
            {
                "project_id": allocation.project_id,
                "project_name": allocation.project_name,
                "client_name": allocation.client_name,
                "period": allocation.period,
                "end_date": _period_end_date(allocation.period),
            }
        )

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    results = []
    bench_count = 0
    partial_count = 0

    for person in enriched_people:
        row = person["row"]
        resolved_employee = person.get("resolved_employee")
        location = person.get("location")
        employee_ref = bench_employee_ref(person)
        designation = _prefer_excel_value(None, getattr(resolved_employee, "designation", None), "-")
        location_code = getattr(location, "code", "")

        if location_filter and location_code != location_filter:
            continue
        if designation_filter and designation.lower() != designation_filter.lower():
            continue
        if utilisation_min is not None and row.utilisation_percent < utilisation_min:
            continue
        if utilisation_max is not None and row.utilisation_percent > utilisation_max:
            continue

        employee_skills = skills_by_emp.get(employee_ref, [])
        if skill_filter:
            skill_names = [skill.skill_name.lower() for skill in employee_skills]
            if skill_filter.lower() not in skill_names:
                continue

        if search:
            search_lower = search.lower()
            if search_lower not in row.employee_name.lower() and search_lower not in designation.lower():
                continue

        if row.classification == "bench":
            bench_count += 1
        elif row.classification == "partially_billed":
            partial_count += 1

        skill_responses = [
            {
                "id": str(skill.id),
                "employee_id": skill.employee_id,
                "skill_name": skill.skill_name,
                "proficiency": skill.proficiency,
                "added_by": skill.added_by,
                "added_at": skill.added_at,
                "notes": skill.notes,
            }
            for skill in employee_skills
        ]

        employee_assignments = assignments_by_emp.get(employee_ref, [])
        active_projects = [
            assignment
            for assignment in employee_assignments
            if assignment["status"] in ("ACTIVE", "ON_HOLD") and not assignment.get("is_deleted")
        ]

        alloc_history = alloc_last_by_emp.get(employee_ref, [])
        if alloc_history:
            last_projects = []
            for alloc_record in alloc_history[:3]:
                last_projects.append(
                    {
                        "project_id": alloc_record["project_id"],
                        "project_name": alloc_record["project_name"],
                        "status": "COMPLETED",
                        "role": employee_project_role.get((employee_ref, alloc_record["project_id"]), ""),
                        "end_date": alloc_record["end_date"],
                        "client_name": alloc_record["client_name"],
                        "period": alloc_record["period"],
                    }
                )
            bench_since = last_projects[0]["end_date"] if last_projects else None
        else:
            last_projects = sorted(
                [assignment for assignment in employee_assignments if assignment["status"] == "COMPLETED" or assignment.get("is_deleted")],
                key=lambda assignment: assignment.get("end_date") or "",
                reverse=True,
            )[:3]
            completed_with_dates = [assignment for assignment in last_projects if assignment.get("end_date")]
            bench_since = completed_with_dates[0]["end_date"] if completed_with_dates else None

        bench_duration_days: Optional[int] = None
        if bench_since:
            try:
                bench_duration_days = (date.today() - date.fromisoformat(bench_since)).days
            except Exception:
                bench_duration_days = None

        available_from = available_from_map.get(employee_ref)
        if row.classification == "bench" and not active_projects:
            available_from = today_str
        elif available_from is None:
            available_from = today_str

        results.append(
            {
                "employee_id": employee_ref,
                "employee_name": row.employee_name,
                "designation": designation,
                "department": _prefer_excel_value(row.department, person.get("department_name"), "Unknown"),
                "location": f"{location.city}, {location.country}" if location else "Excel Import",
                "skills": skill_responses,
                "utilisation_percent": row.utilisation_percent,
                "classification": row.classification,
                "available_from": available_from,
                "current_projects": [{k: v for k, v in project.items() if k != "is_deleted"} for project in active_projects],
                "active_projects": [{k: v for k, v in project.items() if k != "is_deleted"} for project in active_projects],
                "last_projects": [{k: v for k, v in project.items() if k != "is_deleted"} for project in last_projects],
                "bench_since": bench_since,
                "bench_duration_days": bench_duration_days,
            }
        )

    total = len(results)
    bench_durations = [
        result["bench_duration_days"]
        for result in results
        if result["classification"] == "bench" and result["bench_duration_days"] is not None
    ]
    avg_bench_days = round(sum(bench_durations) / len(bench_durations)) if bench_durations else None

    skip = (page - 1) * page_size
    paginated = results[skip : skip + page_size]

    return {
        "period": target_period,
        "data_source": "excel",
        "employees": paginated,
        "total": total,
        "bench_count": bench_count,
        "partial_count": partial_count,
        "avg_bench_days": avg_bench_days,
    }


async def get_excel_designations(
    branch_location_id: str,
    period: Optional[str] = None,
) -> list[str]:
    bench_pool = await get_excel_bench_pool(
        branch_location_id=branch_location_id,
        period=period,
        page=1,
        page_size=5000,
    )
    return sorted({employee["designation"] for employee in bench_pool["employees"] if employee.get("designation")})


async def get_excel_finance_billable(
    period: str,
    branch_location_id: str,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    rows = await ExcelUtilisationReport.find(
        ExcelUtilisationReport.branch_location_id == branch_location_id,
        ExcelUtilisationReport.period == period,
    ).to_list()
    rows = _dedupe_latest_rows(rows)

    if not rows:
        return {
            "entries": [],
            "total": 0,
            "period": period,
            "latest_version": 1,
            "data_source": "excel",
        }

    employee_ids = [row.employee_id for row in rows if row.employee_id]
    allocations = (
        await ProjectAllocation.find(
            ProjectAllocation.period == period,
            {"employee_id": {"$in": employee_ids}},
            ProjectAllocation.is_deleted != True,
        ).to_list()
        if employee_ids
        else []
    )
    allocations_by_employee: dict[str, list[ProjectAllocation]] = defaultdict(list)
    for allocation in allocations:
        allocations_by_employee[allocation.employee_id].append(allocation)
    for employee_allocations in allocations_by_employee.values():
        employee_allocations.sort(key=lambda allocation: allocation.allocation_percentage, reverse=True)

    effective_timesheet_entries = await get_effective_timesheet_entries(
        period=period,
        employee_ids=employee_ids,
    )
    employee_total_hours = summarise_hours_by_employee(effective_timesheet_entries)

    entries = []
    for row in rows:
        employee_id = row.employee_id or _excel_employee_ref(row.employee_name)
        top_allocation = allocations_by_employee.get(row.employee_id, [None])[0]
        derived_billable_hours = _derived_billable_hours(period, row.utilisation_percent, row.classification)
        timesheet_hours = employee_total_hours.get(row.employee_id, {}) if row.employee_id else {}
        billable_hours = round(timesheet_hours.get("billable", derived_billable_hours), 1)
        if row.classification == "bench":
            billable_status = "non_billable"
            billable_hours = 0.0
        else:
            billable_status = row.classification

        entries.append(
            {
                "id": f"excel-finance:{period}:{_normalise_employee_key(row.employee_name, row.employee_id)}",
                "employee_id": employee_id,
                "employee_name": row.employee_name,
                "period": period,
                "billable_status": billable_status,
                "billable_hours": billable_hours,
                "billed_amount": None,
                "project_name": top_allocation.project_name if top_allocation else None,
                "client_name": top_allocation.client_name if top_allocation else None,
                "version": 1,
            }
        )

    total = len(entries)
    skip = (page - 1) * page_size
    return {
        "entries": entries[skip : skip + page_size],
        "total": total,
        "period": period,
        "latest_version": 1,
        "data_source": "excel",
    }


async def get_excel_finance_upload_history(branch_location_id: str) -> list[dict]:
    logs = await ExcelUploadLog.find(
        ExcelUploadLog.branch_location_id == branch_location_id,
    ).sort(-ExcelUploadLog.uploaded_at).limit(20).to_list()

    history = []
    for log in logs:
        periods = log.periods or [None]
        for period in periods:
            history.append(
                {
                    "batch_id": f"{log.batch_id}:{period or 'all'}",
                    "period": period or "Unknown",
                    "uploaded_by": log.uploaded_by,
                    "filename": log.filename,
                    "total_rows": log.total_rows,
                    "valid_count": log.matched_rows,
                    "error_count": max(log.total_rows - log.matched_rows, 0),
                    "duplicate_count": 0,
                    "version": 1,
                    "errors": [],
                    "uploaded_at": log.uploaded_at.isoformat(),
                }
            )

    history.sort(key=lambda item: (item["uploaded_at"], item["period"]), reverse=True)
    return history


async def get_excel_employee_detail(
    employee_ref: str,
    requester_branch_location_id: str,
    period: Optional[str] = None,
) -> Optional[dict]:
    employee_name = _decode_excel_employee_ref(employee_ref)

    # Handle "unmatched:<norm_name>" refs — query by employee_id directly
    if employee_name.startswith("unmatched:"):
        query_filters = [
            ExcelUtilisationReport.branch_location_id == requester_branch_location_id,
            ExcelUtilisationReport.employee_id == employee_name,
        ]
        if period:
            query_filters.append(ExcelUtilisationReport.period == period)
        rows = await ExcelUtilisationReport.find(*query_filters).sort(-ExcelUtilisationReport.uploaded_at).to_list()
        if not rows and period:
            rows = await ExcelUtilisationReport.find(
                ExcelUtilisationReport.branch_location_id == requester_branch_location_id,
                ExcelUtilisationReport.employee_id == employee_name,
            ).sort(-ExcelUtilisationReport.period, -ExcelUtilisationReport.uploaded_at).to_list()
    else:
        query_filters = [ExcelUtilisationReport.branch_location_id == requester_branch_location_id]
        if period:
            query_filters.append(ExcelUtilisationReport.period == period)
        query_filters.append(ExcelUtilisationReport.employee_name == employee_name)

        rows = await ExcelUtilisationReport.find(*query_filters).sort(-ExcelUtilisationReport.uploaded_at).to_list()
        if not rows and period:
            rows = await ExcelUtilisationReport.find(
                ExcelUtilisationReport.branch_location_id == requester_branch_location_id,
                ExcelUtilisationReport.employee_name == employee_name,
            ).sort(-ExcelUtilisationReport.period, -ExcelUtilisationReport.uploaded_at).to_list()

    if not rows:
        return None

    latest_row = rows[0]
    periods = sorted({row.period for row in rows})
    capacity_hours = _capacity_hours_for_period(latest_row.period)
    total_hours = _derived_total_hours(latest_row.period, latest_row.utilisation_percent)
    billable_hours = _derived_billable_hours(latest_row.period, latest_row.utilisation_percent, latest_row.classification)

    # Check if this employee belongs to another branch
    stored_id = latest_row.employee_id
    is_other_branch = False
    real_employee: Optional[Employee] = None
    if stored_id and not stored_id.startswith("unmatched:") and ObjectId.is_valid(stored_id):
        real_employee = await Employee.get(stored_id)
        if real_employee and real_employee.location_id != requester_branch_location_id:
            is_other_branch = True

    if is_other_branch and real_employee:
        loc = await Location.get(real_employee.location_id)
        dept = await Department.get(real_employee.department_id)
        return {
            "id": str(real_employee.id),
            "name": real_employee.name,
            "designation": real_employee.designation,
            "department": dept.name if dept else "Unknown",
            "department_id": real_employee.department_id,
            "level": real_employee.level,
            "location_id": real_employee.location_id,
            "location_code": loc.code if loc else "UNK",
            "location_city": loc.city if loc else "Unknown",
            "photo_url": real_employee.photo_url,
            "is_active": real_employee.is_active,
            "join_date": None,
            "tenure_months": None,
            "managers": [],
            "reporting_chain": [],
            "direct_reports": [],
            "projects": [],
            "skills": [],
            "is_own_branch": False,
            "is_other_branch": True,
            "data_source": "excel",
        }

    return {
        "id": employee_ref,
        "name": latest_row.employee_name,
        "email": latest_row.employee_email,
        "designation": latest_row.department or "YTPL Employee",
        "department": latest_row.department or "Unknown",
        "department_id": "",
        "level": "unknown",
        "location_id": requester_branch_location_id,
        "location_code": "YTPL",
        "location_city": "Not in system",
        "photo_url": None,
        "is_active": True,
        "join_date": None,
        "tenure_months": None,
        "managers": [],
        "reporting_chain": [],
        "direct_reports": [],
        "projects": [],
        "skills": [],
        "utilisation": {
            "period": latest_row.period,
            "utilisation_percent": latest_row.utilisation_percent,
            "billable_percent": latest_row.utilisation_percent if latest_row.classification != "bench" else 0.0,
            "total_hours": total_hours,
            "billable_hours": billable_hours,
            "non_billable_hours": round(total_hours - billable_hours, 1),
            "capacity_hours": capacity_hours,
            "classification": latest_row.classification,
        },
        "timesheet_summary": {
            "period": latest_row.period,
            "total_hours": total_hours,
            "billable_hours": billable_hours,
            "entry_count": 1,
        },
        "is_own_branch": True,
        "is_unmatched": True,
        "excel_periods": periods,
        "data_source": "excel",
    }


async def get_excel_overlay_for_employee(
    *,
    employee_id: Optional[str],
    employee_name: Optional[str],
    branch_location_id: str,
    period: Optional[str],
) -> Optional[dict]:
    filters: list = [ExcelUtilisationReport.branch_location_id == branch_location_id]
    if period:
        filters.append(ExcelUtilisationReport.period == period)
    if employee_id:
        filters.append(ExcelUtilisationReport.employee_id == employee_id)
    elif employee_name:
        filters.append(ExcelUtilisationReport.employee_name == employee_name)
    else:
        return None

    rows = await ExcelUtilisationReport.find(*filters).sort(-ExcelUtilisationReport.uploaded_at).to_list()
    if not rows and period and employee_name:
        rows = await ExcelUtilisationReport.find(
            ExcelUtilisationReport.branch_location_id == branch_location_id,
            ExcelUtilisationReport.employee_name == employee_name,
        ).sort(-ExcelUtilisationReport.period, -ExcelUtilisationReport.uploaded_at).to_list()

    if not rows:
        return None

    latest_row = rows[0]
    total_hours = _derived_total_hours(latest_row.period, latest_row.utilisation_percent)
    billable_hours = _derived_billable_hours(latest_row.period, latest_row.utilisation_percent, latest_row.classification)

    return {
        "period": latest_row.period,
        "utilisation_percent": latest_row.utilisation_percent,
        "billable_percent": latest_row.utilisation_percent if latest_row.classification != "bench" else 0.0,
        "total_hours": total_hours,
        "billable_hours": billable_hours,
        "non_billable_hours": round(total_hours - billable_hours, 1),
        "capacity_hours": _capacity_hours_for_period(latest_row.period),
        "classification": latest_row.classification,
        "timesheet_summary": {
            "period": latest_row.period,
            "total_hours": total_hours,
            "billable_hours": billable_hours,
            "entry_count": 1,
        },
    }


async def get_upload_history(branch_location_id: str) -> list[dict]:
    logs = await ExcelUploadLog.find(
        ExcelUploadLog.branch_location_id == branch_location_id,
    ).sort(-ExcelUploadLog.uploaded_at).limit(20).to_list()

    return [
        {
            "batch_id": log.batch_id,
            "filename": log.filename,
            "total_rows": log.total_rows,
            "matched_rows": log.matched_rows,
            "periods": log.periods,
            "uploaded_at": log.uploaded_at.isoformat(),
        }
        for log in logs
    ]


def _normalise_project_lookup_key(
    project_name: Optional[str],
    client_name: Optional[str],
) -> tuple[str, str]:
    return (_normalise_name(project_name or ""), _normalise_name(client_name or ""))


def _build_project_planned_worked_lookups(records) -> tuple[dict[tuple[str, str], dict], dict[str, dict]]:
    exact_lookup: dict[tuple[str, str], dict] = {}
    by_name_lookup: dict[str, dict] = {}
    for rec in records:
        exact_key = _normalise_project_lookup_key(rec.project_name, rec.client_name)
        name_key = exact_key[0]

        exact_bucket = exact_lookup.setdefault(
            exact_key,
            {"planned_days": 0.0, "worked_days": 0.0, "record_count": 0},
        )
        exact_bucket["planned_days"] += rec.planned_days
        exact_bucket["worked_days"] += rec.worked_days
        exact_bucket["record_count"] += 1

        name_bucket = by_name_lookup.setdefault(
            name_key,
            {"planned_days": 0.0, "worked_days": 0.0, "record_count": 0},
        )
        name_bucket["planned_days"] += rec.planned_days
        name_bucket["worked_days"] += rec.worked_days
        name_bucket["record_count"] += 1

    return exact_lookup, by_name_lookup


def _resolve_project_planned_worked(
    project_name: Optional[str],
    client_name: Optional[str],
    exact_lookup: dict[tuple[str, str], dict],
    by_name_lookup: dict[str, dict],
) -> Optional[dict]:
    exact_key = _normalise_project_lookup_key(project_name, client_name)
    summary = exact_lookup.get(exact_key)
    if summary and summary["record_count"] > 0:
        return summary

    name_key = exact_key[0]
    summary = by_name_lookup.get(name_key)
    if summary and summary["record_count"] > 0:
        return summary

    return None


def _finalise_excel_project_progress(
    planned_worked_summary: Optional[dict],
    member_count: int,
    total_allocated_days: float,
) -> tuple[float, float, float]:
    if planned_worked_summary and planned_worked_summary["record_count"] > 0:
        planned_days = round(planned_worked_summary["planned_days"], 1)
        worked_days = round(planned_worked_summary["worked_days"], 1)
        progress_percent = (
            round(min(100.0, worked_days / planned_days * 100), 1)
            if planned_days > 0
            else 0.0
        )
        return planned_days, worked_days, progress_percent

    capacity = member_count * 21.0
    worked_days = round(total_allocated_days, 1)
    planned_days = round(capacity, 1)
    progress_percent = (
        round(min(100.0, total_allocated_days / capacity * 100), 1)
        if capacity
        else 0.0
    )
    return planned_days, worked_days, progress_percent


async def get_excel_project_detail(
    project_id: str,
    period: str,
    branch_location_id: str,
) -> Optional[dict]:
    """
    Return a ProjectDetail-shaped dict for an inter-company Excel project for the selected month.
    """

    # Fetch the project
    project = await Project.find_one(
        {"_id": ObjectId(project_id)} if ObjectId.is_valid(project_id) else {"_id": project_id},
        Project.is_deleted != True,
    )
    if not project:
        return None

    allocations = await ProjectAllocation.find(
        ProjectAllocation.project_id == project_id,
        ProjectAllocation.period == period,
        ProjectAllocation.source_system == "intercompany_excel",
        ProjectAllocation.is_deleted != True,
    ).to_list()

    if not allocations:
        return None

    # Load matched employees for extra info
    allocation_emp_ids = list({a.employee_id for a in allocations})
    real_emp_ids = [
        ObjectId(a.employee_id)
        for a in allocations
        if ObjectId.is_valid(a.employee_id) and not a.employee_id.startswith("unmatched:")
    ]
    employees_by_id: dict[str, Employee] = {}
    if real_emp_ids:
        emps = await Employee.find({"_id": {"$in": real_emp_ids}}).to_list()
        employees_by_id = {str(e.id): e for e in emps}

    # Reporting relationships for line managers
    emp_id_strs = list(employees_by_id.keys())
    reporting_rels = (
        await ReportingRelationship.find(
            {"employee_id": {"$in": emp_id_strs}},
            ReportingRelationship.type == "PRIMARY",
            ReportingRelationship.is_deleted != True,
        ).to_list()
        if emp_id_strs
        else []
    )
    manager_ids = [r.manager_id for r in reporting_rels]
    managers = (
        await Employee.find(
            {"_id": {"$in": [ObjectId(mid) for mid in manager_ids if ObjectId.is_valid(mid)]}}
        ).to_list()
        if manager_ids
        else []
    )
    manager_name_by_id = {str(m.id): m.name for m in managers}
    line_manager_by_emp: dict[str, str] = {
        r.employee_id: manager_name_by_id.get(r.manager_id, "No Manager")
        for r in reporting_rels
    }

    members = []
    total_allocated_days = 0.0

    # Load planned/worked days for this project and period
    from app.models.employee_planned_worked import EmployeePlannedWorked
    all_pw_records = (
        await EmployeePlannedWorked.find(
            {"employee_id": {"$in": allocation_emp_ids}, "period": period},
        ).to_list()
        if allocation_emp_ids
        else []
    )
    exact_pw_lookup, name_pw_lookup = _build_project_planned_worked_lookups(all_pw_records)
    project_pw_summary = _resolve_project_planned_worked(
        project.name,
        project.client_name,
        exact_pw_lookup,
        name_pw_lookup,
    )

    target_exact_key = _normalise_project_lookup_key(project.name, project.client_name)
    has_exact_project_pw = target_exact_key in exact_pw_lookup
    pw_by_emp: dict[str, dict] = {}
    for pw in all_pw_records:
        pw_exact_key = _normalise_project_lookup_key(pw.project_name, pw.client_name)
        if has_exact_project_pw:
            if pw_exact_key != target_exact_key:
                continue
        elif pw_exact_key[0] != target_exact_key[0]:
            continue

        if pw.employee_id not in pw_by_emp:
            pw_by_emp[pw.employee_id] = {"planned_days": 0.0, "worked_days": 0.0}
        pw_by_emp[pw.employee_id]["planned_days"] += pw.planned_days
        pw_by_emp[pw.employee_id]["worked_days"] += pw.worked_days

    for alloc in allocations:
        emp = employees_by_id.get(alloc.employee_id)
        total_allocated_days += alloc.allocated_days
        pw = pw_by_emp.get(alloc.employee_id)
        members.append({
            "employee_id": alloc.employee_id,
            "employee_name": alloc.employee_name,
            "line_manager": line_manager_by_emp.get(alloc.employee_id, "No Manager"),
            "designation": emp.designation if emp else "Employee",
            "department": "",
            "location": "",
            "location_code": "",
            "role_in_project": "Consultant",
            "assigned_at": None,
            "allocation_percentage": alloc.allocation_percentage,
            "allocated_days": pw["planned_days"] if pw else alloc.allocated_days,
            "worked_days": pw["worked_days"] if pw else None,
        })

    # Sort by allocation % desc
    members.sort(key=lambda m: m["allocation_percentage"] or 0, reverse=True)

    member_count = len(members)
    avg_alloc_pct = round(
        sum(m["allocation_percentage"] or 0 for m in members) / member_count, 1
    ) if member_count else 0.0

    total_planned, total_worked, progress_percent = _finalise_excel_project_progress(
        project_pw_summary,
        member_count,
        total_allocated_days,
    )
    if not project_pw_summary and total_planned > 0:
        progress_percent = avg_alloc_pct if avg_alloc_pct > 0 else progress_percent

    return {
        "id": str(project.id),
        "name": project.name,
        "status": project.status,
        "project_type": project.project_type,
        "client_name": project.client_name or "",
        "description": project.description,
        "start_date": _period_date(period),
        "end_date": _period_end_date(period),
        "planned_days": total_planned,
        "worked_days": total_worked,
        "progress_percent": progress_percent,
        "member_count": member_count,
        "members": members,
        "period": period,
        "data_source": "excel",
    }


async def get_excel_projects(
    branch_location_id: str,
    period: str,
    search: Optional[str] = None,
    client_name: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    """
    Return project list from inter-company Excel allocations for the selected month.
    Shape mirrors the HRMS listProjects response so the frontend table works unchanged.
    """
    allocations = await ProjectAllocation.find(
        ProjectAllocation.period == period,
        ProjectAllocation.source_system == "intercompany_excel",
        ProjectAllocation.is_deleted != True,
    ).to_list()

    allocation_emp_ids = list({alloc.employee_id for alloc in allocations})
    from app.models.employee_planned_worked import EmployeePlannedWorked
    pw_records = (
        await EmployeePlannedWorked.find(
            {"employee_id": {"$in": allocation_emp_ids}, "period": period},
        ).to_list()
        if allocation_emp_ids
        else []
    )
    exact_pw_lookup, name_pw_lookup = _build_project_planned_worked_lookups(pw_records)

    # Group by project_id
    project_map: dict[str, dict] = {}
    for alloc in allocations:
        pid = alloc.project_id
        if pid not in project_map:
            project_map[pid] = {
                "id": pid,
                "name": alloc.project_name,
                "client_name": alloc.client_name or "",
                "status": "ACTIVE",
                "project_type": "client",
                "member_count": 0,
                "total_allocated_days": 0.0,
                "start_date": _period_date(period),
                "end_date": _period_end_date(period),
                "planned_days": 0,
                "worked_days": 0,
                "progress_percent": 0.0,
            }
        project_map[pid]["member_count"] += 1
        project_map[pid]["total_allocated_days"] += alloc.allocated_days

    projects = list(project_map.values())

    # Filters
    if search:
        sl = search.lower()
        projects = [p for p in projects if sl in p["name"].lower() or sl in (p["client_name"] or "").lower()]
    if client_name:
        projects = [p for p in projects if p["client_name"] == client_name]

    # Sort by member count desc
    projects.sort(key=lambda p: p["member_count"], reverse=True)

    # Prefer planned/worked sheet totals; only fall back to allocation-derived capacity if the month has no Excel planned/worked rows.
    for p in projects:
        summary = _resolve_project_planned_worked(
            p["name"],
            p["client_name"],
            exact_pw_lookup,
            name_pw_lookup,
        )
        planned_days, worked_days, progress_percent = _finalise_excel_project_progress(
            summary,
            p["member_count"],
            p["total_allocated_days"],
        )
        p["planned_days"] = planned_days
        p["worked_days"] = worked_days
        p["progress_percent"] = progress_percent

    total = len(projects)
    start = (page - 1) * page_size
    page_slice = projects[start: start + page_size]

    # Unique clients for filter options
    clients = sorted({p["client_name"] for p in projects if p["client_name"]})

    return {
        "projects": page_slice,
        "total": total,
        "active_count": total,
        "completed_count": 0,
        "on_hold_count": 0,
        "period": period,
        "data_source": "excel",
        "clients": clients,
    }


async def _reimport_intercompany(file_path, branch_location_id: str) -> dict:
    """
    Re-run the inter-company sheet import inline (mirrors import_intercompany_excel seed).
    Used by the /reimport endpoint triggered from the UI toggle.
    """
    import io as _io
    import re as _re
    from pathlib import Path as _Path

    try:
        import openpyxl as _openpyxl
    except ImportError:
        return {"error": "openpyxl not installed"}

    PERIOD = "2026-03"
    TARGET_SHEET = "Inter-company"
    COMPANY_FILTER = {"YTPL"}
    WORKING_DAYS = 21
    HOURS_PER_DAY = 8.0
    CAPACITY_HOURS = WORKING_DAYS * HOURS_PER_DAY

    def _norm(v: str) -> str:
        c = v.lower().strip()
        c = _re.sub(r"\s*[-(]?\s*(onsite|offshore)\s*[)]?\s*$", "", c)
        c = _re.sub(r"[^a-z0-9]+", "", c)
        return c

    def _parse_hours(v) -> float:
        if v is None:
            return 0.0
        try:
            return max(0.0, float(v))
        except (TypeError, ValueError):
            return 0.0

    wb = _openpyxl.load_workbook(_io.BytesIO(_Path(file_path).read_bytes()), data_only=True)
    if TARGET_SHEET not in wb.sheetnames:
        return {"error": f"Sheet '{TARGET_SHEET}' not found"}

    ws = wb[TARGET_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row with datetime month columns
    header_row_idx, header, march_col = None, None, None
    for idx, row in enumerate(rows[:10]):
        for i, cell in enumerate(row or []):
            if isinstance(cell, datetime) and cell.year == 2026 and cell.month == 3:
                header_row_idx, header, march_col = idx, row, i
                break
        if march_col is not None:
            break

    if march_col is None:
        return {"error": "March 2026 column not found"}

    # Parse records
    records = []
    for row in rows[header_row_idx + 1:]:
        if not row or len(row) <= march_col:
            continue
        name = str(row[1]).strip() if row[1] else ""
        company = str(row[2]).strip() if row[2] else ""
        client = str(row[4]).strip() if row[4] else ""
        project = str(row[6]).strip() if row[6] else ""
        if not name or not company or company.upper() not in COMPANY_FILTER:
            continue
        nl = name.lower().strip()
        if nl.startswith(("tbc", "total", "sub-total")):
            continue
        hours = _parse_hours(row[march_col])
        if hours <= 0:
            continue
        records.append({"name": name, "client": client, "project": project, "hours": hours})

    # Load employees for this branch
    employees = await Employee.find(
        Employee.location_id == branch_location_id,
        Employee.is_active == True,
    ).to_list()
    exact_map = {e.name.strip().lower(): e for e in employees}
    norm_map = {_norm(e.name): e for e in employees}

    now = datetime.now(timezone.utc)

    # Upsert projects
    project_cache: dict[tuple, Project] = {}
    async def _get_proj(client: str, proj_name: str) -> Project:
        key = (client, proj_name)
        if key not in project_cache:
            existing = await Project.find_one(
                Project.name == proj_name, Project.client_name == client, Project.is_deleted != True
            ) or await Project.find_one(Project.name == proj_name, Project.is_deleted != True)
            if existing:
                if not existing.client_name:
                    existing.client_name = client
                    existing.updated_at = now
                    await existing.save()
                project_cache[key] = existing
            else:
                p = Project(
                    name=proj_name, client_name=client, status="ACTIVE", project_type="client",
                    start_date=datetime(2026, 3, 1, tzinfo=timezone.utc), created_at=now, updated_at=now,
                )
                await p.insert()
                project_cache[key] = p
        return project_cache[key]

    # Delete existing intercompany allocations for this period
    await ProjectAllocation.find(
        ProjectAllocation.period == PERIOD,
        ProjectAllocation.source_system == "intercompany_excel",
    ).delete()

    # Pass 1: upsert all projects
    for rec in records:
        await _get_proj(rec["client"], rec["project"])

    # Pass 2: store allocations
    matched = unmatched = inserted = 0
    for rec in records:
        proj = await _get_proj(rec["client"], rec["project"])
        proj_id = str(proj.id)
        emp = exact_map.get(rec["name"].strip().lower()) or norm_map.get(_norm(rec["name"]))
        hours = rec["hours"]
        alloc_pct = round(min(100.0, hours / CAPACITY_HOURS * 100), 1)
        avail_days = round(max(0.0, WORKING_DAYS - hours / HOURS_PER_DAY), 1)

        if emp:
            emp_id = str(emp.id)
            matched += 1
        else:
            emp_id = f"unmatched:{_norm(rec['name'])}"
            unmatched += 1

        await ProjectAllocation(
            employee_id=emp_id,
            hrms_employee_id=getattr(emp, "hrms_employee_id", 0) or 0 if emp else 0,
            employee_name=rec["name"] if not emp else emp.name,
            project_id=proj_id,
            hrms_project_id=getattr(proj, "hrms_project_id", 0) or 0,
            project_name=proj.name,
            client_name=proj.client_name,
            period=PERIOD,
            allocated_days=round(hours / HOURS_PER_DAY, 1),
            allocation_percentage=alloc_pct,
            total_working_days=WORKING_DAYS,
            available_days=avail_days,
            source_system="intercompany_excel",
            source_id=f"intercompany:{PERIOD}:{emp_id}:{proj_id}",
            created_at=now,
            updated_at=now,
        ).insert()
        inserted += 1

        if emp:
            exists = await EmployeeProject.find_one(
                EmployeeProject.employee_id == emp_id,
                EmployeeProject.project_id == proj_id,
            )
            if not exists:
                await EmployeeProject(
                    employee_id=emp_id, project_id=proj_id, role_in_project="Consultant",
                    start_date=datetime(2026, 3, 1, tzinfo=timezone.utc), created_at=now, updated_at=now,
                ).insert()

    return {
        "projects": len(project_cache),
        "allocations": inserted,
        "matched_employees": matched,
        "unmatched_employees": unmatched,
    }


async def run_configured_excel_reimport(branch_location_id: str, user_id: str) -> dict:
    """Re-run the configured Excel import flow used by the UI and integration sync."""
    from seed.seed_excel import import_planned_worked, update_workbook_employee_split

    file_path = Path(settings.EXCEL_FILE_PATH)
    if not file_path.exists():
        return {
            "status": "skipped",
            "reason": f"Excel file not found at configured path: {settings.EXCEL_FILE_PATH}",
            "file": str(file_path),
        }

    content = file_path.read_bytes()
    filename = file_path.name
    now = datetime.now(timezone.utc)

    ytpl_result = await parse_and_store_excel(
        file_content=content,
        filename=filename,
        branch_location_id=branch_location_id,
        user_id=user_id,
    )
    if "error" in ytpl_result:
        return {
            "status": "failed",
            "reason": ytpl_result["error"],
            "file": str(file_path),
        }

    intercompany_result = await _reimport_intercompany(file_path, branch_location_id)

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    planned_worked_result = await import_planned_worked(workbook, branch_location_id, filename, now)
    workbook_update_path = await update_workbook_employee_split(file_path)

    return {
        "status": "completed",
        "file": str(workbook_update_path),
        "ytpl": ytpl_result,
        "intercompany": intercompany_result,
        "planned_worked": planned_worked_result,
    }


async def get_employee_planned_worked_timeline(employee_id: str) -> dict:
    """
    Return month-by-month planned vs worked days for an employee from the Excel sheets.
    Aggregates across all projects per period.
    """
    from app.models.employee_planned_worked import EmployeePlannedWorked

    if employee_id.startswith(EXCEL_EMPLOYEE_PREFIX):
        employee_name = _decode_excel_employee_ref(employee_id)
        records = await EmployeePlannedWorked.find(
            EmployeePlannedWorked.employee_name == employee_name,
        ).to_list()
        utilisation_rows = await ExcelUtilisationReport.find(
            ExcelUtilisationReport.employee_name == employee_name,
        ).to_list()
    else:
        records = await EmployeePlannedWorked.find(
            EmployeePlannedWorked.employee_id == employee_id,
        ).to_list()
        utilisation_rows = await ExcelUtilisationReport.find(
            ExcelUtilisationReport.employee_id == employee_id,
        ).to_list()

    if not records:
        # Try by unmatched key — shouldn't happen for real employees but fallback
        return {"employee_id": employee_id, "timeline": []}

    # Aggregate per period
    period_data: dict[str, dict] = {}
    for rec in records:
        p = rec.period
        if p not in period_data:
            period_data[p] = {"planned_days": 0.0, "worked_days": 0.0, "projects": []}
        period_data[p]["planned_days"] = round(period_data[p]["planned_days"] + rec.planned_days, 1)
        period_data[p]["worked_days"] = round(period_data[p]["worked_days"] + rec.worked_days, 1)
        if rec.project_name:
            period_data[p]["projects"].append({
                "project_name": rec.project_name,
                "client_name": rec.client_name,
                "planned_days": rec.planned_days,
                "worked_days": rec.worked_days,
            })

    timeline = [
        {
            "period": period,
            "planned_days": data["planned_days"],
            "worked_days": data["worked_days"],
            "utilisation_percent": round(
                min(100.0, data["worked_days"] / data["planned_days"] * 100), 1
            ) if data["planned_days"] > 0 else 0.0,
            "projects": sorted(data["projects"], key=lambda x: x["planned_days"], reverse=True),
        }
        for period, data in sorted(period_data.items())
    ]

    return {"employee_id": employee_id, "timeline": timeline}


async def get_branch_planned_vs_actual(period: str, branch_location_id: str) -> Optional[dict]:
    """
    Return branch-level planned vs actual utilisation for a given period.
    Aggregates EmployeePlannedWorked records for employees in this branch.
    """
    from app.models.employee_planned_worked import EmployeePlannedWorked

    # Get employee IDs for this branch
    employees = await Employee.find(
        Employee.location_id == branch_location_id,
        Employee.is_active == True,
    ).to_list()
    emp_ids = [str(e.id) for e in employees]
    if not emp_ids:
        return None

    records = await EmployeePlannedWorked.find(
        {"employee_id": {"$in": emp_ids}, "period": period}
    ).to_list()

    if not records:
        return None

    total_planned = round(sum(r.planned_days for r in records), 1)
    total_worked = round(sum(r.worked_days for r in records), 1)
    unique_employees = len({r.employee_id for r in records})

    # Per-project breakdown
    proj_map: dict[str, dict] = {}
    for r in records:
        key = r.project_name or "Unassigned"
        if key not in proj_map:
            proj_map[key] = {"project_name": key, "client_name": r.client_name, "planned_days": 0.0, "worked_days": 0.0}
        proj_map[key]["planned_days"] = round(proj_map[key]["planned_days"] + r.planned_days, 1)
        proj_map[key]["worked_days"] = round(proj_map[key]["worked_days"] + r.worked_days, 1)

    projects = sorted(proj_map.values(), key=lambda x: x["planned_days"], reverse=True)

    planned_util_pct = round(total_planned / (unique_employees * 21) * 100, 1) if unique_employees else 0.0
    actual_util_pct = round(total_worked / (unique_employees * 21) * 100, 1) if unique_employees else 0.0

    return {
        "period": period,
        "total_planned_days": total_planned,
        "total_worked_days": total_worked,
        "employee_count": unique_employees,
        "planned_utilisation_pct": planned_util_pct,
        "actual_utilisation_pct": actual_util_pct,
        "variance_pct": round(actual_util_pct - planned_util_pct, 1),
        "projects": projects[:10],
    }


async def list_excel_employees(
    branch_location_id: str,
    period: str = "2026-03",
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    """Return unmatched Excel employees (not in HRMS) for the Employee Master page."""
    rows = await ExcelUtilisationReport.find(
        ExcelUtilisationReport.branch_location_id == branch_location_id,
        ExcelUtilisationReport.period == period,
    ).to_list()

    rows = _dedupe_latest_rows(rows)

    # Only unmatched employees
    unmatched = [r for r in rows if r.employee_id and r.employee_id.startswith("unmatched:")]

    if search:
        sl = search.lower()
        unmatched = [r for r in unmatched if sl in r.employee_name.lower() or sl in (r.department or "").lower()]

    total = len(unmatched)
    start = (page - 1) * page_size
    page_rows = unmatched[start: start + page_size]

    employees = []
    for r in page_rows:
        employees.append({
            "id": r.employee_id,
            "name": r.employee_name,
            "email": r.employee_email or "-",
            "designation": "-",
            "department": r.department or "Unknown",
            "level": "-",
            "location": "Excel Import",
            "join_date": None,
            "is_active": True,
            "classification": r.classification,
            "utilisation_percent": r.utilisation_percent,
            "availability_percent": r.availability_percent,
        })

    return {
        "employees": employees,
        "total": total,
        "active_count": total,
        "inactive_count": 0,
    }


async def get_employee_planned_worked_timeline(employee_id: str) -> dict:
    """
    Return month-by-month planned vs worked days for an employee from the Excel sheets.
    Missing Excel values fall back to 20 days so the drawer never renders blank months as zero.
    """
    from app.models.employee_planned_worked import EmployeePlannedWorked

    if employee_id.startswith(EXCEL_EMPLOYEE_PREFIX):
        employee_name = _decode_excel_employee_ref(employee_id)
        records = await EmployeePlannedWorked.find(
            EmployeePlannedWorked.employee_name == employee_name,
        ).to_list()
        utilisation_rows = await ExcelUtilisationReport.find(
            ExcelUtilisationReport.employee_name == employee_name,
        ).to_list()
    else:
        records = await EmployeePlannedWorked.find(
            EmployeePlannedWorked.employee_id == employee_id,
        ).to_list()
        utilisation_rows = await ExcelUtilisationReport.find(
            ExcelUtilisationReport.employee_id == employee_id,
        ).to_list()

    if not records and not utilisation_rows:
        return {"employee_id": employee_id, "timeline": []}

    period_data: dict[str, dict] = {}
    for rec in records:
        period_bucket = period_data.setdefault(
            rec.period,
            {"planned_days": 0.0, "worked_days": 0.0, "projects": []},
        )
        period_bucket["planned_days"] = round(period_bucket["planned_days"] + rec.planned_days, 1)
        period_bucket["worked_days"] = round(period_bucket["worked_days"] + rec.worked_days, 1)
        if rec.project_name:
            period_bucket["projects"].append(
                {
                    "project_name": rec.project_name,
                    "client_name": rec.client_name,
                    "planned_days": rec.planned_days,
                    "worked_days": rec.worked_days,
                }
            )

    for row in _dedupe_latest_rows(utilisation_rows):
        period_data.setdefault(row.period, {"planned_days": 0.0, "worked_days": 0.0, "projects": []})

    timeline = []
    for period, data in sorted(period_data.items()):
        planned_days = round(data["planned_days"], 1) if data["planned_days"] > 0 else DEFAULT_PLANNED_WORKED_DAYS
        worked_days = round(data["worked_days"], 1) if data["worked_days"] > 0 else DEFAULT_PLANNED_WORKED_DAYS
        timeline.append(
            {
                "period": period,
                "planned_days": planned_days,
                "worked_days": worked_days,
                "utilisation_percent": round(min(100.0, worked_days / planned_days * 100), 1) if planned_days > 0 else 0.0,
                "projects": sorted(data["projects"], key=lambda x: x["planned_days"], reverse=True),
            }
        )

    return {"employee_id": employee_id, "timeline": timeline}


async def list_excel_employees(
    branch_location_id: str,
    period: str = "2026-03",
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    """
    Return Excel employees for the Employee Master page.
    Excel remains the source of truth; HRMS only fills fields that Excel does not provide.
    """
    rows = await _get_combined_excel_rows_for_period(period, branch_location_id)
    enriched_people = await _enrich_excel_people(rows, branch_location_id)
    all_enriched_people = await _enrich_excel_people(
        await _get_all_excel_employee_rows(branch_location_id),
        branch_location_id,
    )

    if search:
        search_lower = search.lower()
        enriched_people = [
            person
            for person in enriched_people
            if search_lower in person["row"].employee_name.lower()
            or search_lower in str(_prefer_excel_value(person["row"].employee_email, getattr(person.get("resolved_employee"), "email", None), "")).lower()
            or search_lower in str(_prefer_excel_value(None, getattr(person.get("resolved_employee"), "designation", None), "")).lower()
            or search_lower in str(_prefer_excel_value(person["row"].department, person.get("department_name"), "")).lower()
            or search_lower in str(_prefer_excel_value(None, getattr(person.get("resolved_employee"), "level", None), "")).lower()
        ]

    total = len(enriched_people)
    active_count = sum(1 for person in enriched_people if getattr(person.get("resolved_employee"), "is_active", True))
    inactive_count = total - active_count
    overall_total, overall_active_count, overall_inactive_count = _summarise_excel_people(
        all_enriched_people
    )

    start = (page - 1) * page_size
    page_rows = enriched_people[start: start + page_size]

    employees = []
    for person in page_rows:
        row = person["row"]
        resolved_employee = person.get("resolved_employee")
        location = person.get("location")
        employee_ref = person["resolved_employee_id"] or row.employee_id or _excel_employee_ref(row.employee_name)
        employees.append(
            {
                "id": employee_ref,
                "name": row.employee_name,
                "email": _prefer_excel_value(row.employee_email, getattr(resolved_employee, "email", None), "-"),
                "designation": _prefer_excel_value(None, getattr(resolved_employee, "designation", None), "-"),
                "department": _prefer_excel_value(row.department, person.get("department_name"), "Unknown"),
                "level": _prefer_excel_value(None, getattr(resolved_employee, "level", None), "-"),
                "location": f"{location.city}, {location.country}" if location else "Excel Import",
                "join_date": getattr(resolved_employee, "join_date", None).isoformat() if getattr(resolved_employee, "join_date", None) else None,
                "is_active": getattr(resolved_employee, "is_active", True),
                "classification": row.classification,
                "utilisation_percent": row.utilisation_percent,
                "availability_percent": row.availability_percent,
            }
        )

    return {
        "employees": employees,
        "total": total,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "overall_total": overall_total,
        "overall_active_count": overall_active_count,
        "overall_inactive_count": overall_inactive_count,
    }
