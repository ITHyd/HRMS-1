"""
Import Inter-Company sheet from the Utilisation Report Excel.

Reads the 'Inter-company' sheet, filters YTPL rows, extracts March 2026
hours per employee per project, then:
  1. Upserts Project records (with client_name)
  2. Upserts ProjectAllocation records for 2026-03
  3. Upserts EmployeeProject assignments

Usage:
    cd backend
    python -m seed.import_intercompany_excel ^
        --file "C:\\Users\\sahit\\Downloads\\Utilisation Report 6th March 2026.xlsx" ^
        --user-email manager@nxzen.com
"""

import argparse
import asyncio
import io
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from app.database import init_db
from app.models.employee import Employee
from app.models.employee_project import EmployeeProject
from app.models.project import Project
from app.models.project_allocation import ProjectAllocation
from app.models.user import User

PERIOD = "2026-03"
TARGET_SHEET = "Inter-company"
COMPANY_FILTER = {"YTPL"}
WORKING_DAYS = 21  # March 2026 working days
HOURS_PER_DAY = 8.0
CAPACITY_HOURS = WORKING_DAYS * HOURS_PER_DAY  # 168


def _normalise_name(value: str) -> str:
    cleaned = value.lower().strip()
    cleaned = re.sub(r"\s*[-(]?\s*(onsite|offshore)\s*[)]?\s*$", "", cleaned)
    cleaned = re.sub(r"[^a-z0-9]+", "", cleaned)
    return cleaned


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower().strip()).strip("-")


def _is_person_row(name: str, company: str) -> bool:
    if not name or not company:
        return False
    if company.upper() not in COMPANY_FILTER:
        return False
    name_lower = name.lower().strip()
    if name_lower.startswith("tbc"):
        return False
    if name_lower.startswith("total"):
        return False
    if name_lower.startswith("sub-total"):
        return False
    return True


def _parse_hours(value) -> float:
    if value is None:
        return 0.0
    try:
        h = float(value)
        return max(0.0, h)
    except (TypeError, ValueError):
        return 0.0


def _find_march_col(header: tuple) -> Optional[int]:
    for i, cell in enumerate(header):
        if isinstance(cell, datetime) and cell.year == 2026 and cell.month == 3:
            return i
    return None


def _parse_sheet(ws) -> list[dict]:
    """Return list of {name, client, project, location, hours} for YTPL March rows."""
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (has datetime month columns)
    header_row_idx = None
    header = None
    for idx, row in enumerate(rows[:10]):
        if any(isinstance(cell, datetime) for cell in (row or [])):
            header_row_idx = idx
            header = row
            break

    if header is None:
        raise ValueError("Could not find header row with month columns in Inter-company sheet")

    march_col = _find_march_col(header)
    if march_col is None:
        raise ValueError("March 2026 column not found in Inter-company sheet")

    records = []
    for row in rows[header_row_idx + 1:]:
        if not row or len(row) <= march_col:
            continue
        name = str(row[1]).strip() if row[1] else ""
        company = str(row[2]).strip() if row[2] else ""
        client = str(row[4]).strip() if row[4] else ""
        project = str(row[6]).strip() if row[6] else ""
        location = str(row[7]).strip() if row[7] else ""

        if not _is_person_row(name, company):
            continue

        hours = _parse_hours(row[march_col])
        if hours <= 0:
            continue

        records.append({
            "name": name,
            "client": client,
            "project": project,
            "location": location,
            "hours": hours,
        })

    return records


def _build_employee_maps(employees: list[Employee]):
    exact = {}
    normalised = {}
    for emp in employees:
        exact.setdefault(emp.name.strip().lower(), emp)
        normalised.setdefault(_normalise_name(emp.name), emp)
    return exact, normalised


def _resolve_employee(name: str, exact, normalised) -> Optional[Employee]:
    return exact.get(name.strip().lower()) or normalised.get(_normalise_name(name))


async def _upsert_project(client_name: str, project_name: str, now: datetime) -> Project:
    """Find or create a Project by name + client."""
    existing = await Project.find_one(
        Project.name == project_name,
        Project.client_name == client_name,
        Project.is_deleted != True,
    )
    if existing:
        return existing

    # Also try by name only
    existing = await Project.find_one(
        Project.name == project_name,
        Project.is_deleted != True,
    )
    if existing:
        if not existing.client_name:
            existing.client_name = client_name
            existing.updated_at = now
            await existing.save()
        return existing

    proj = Project(
        name=project_name,
        client_name=client_name,
        status="ACTIVE",
        project_type="client",
        start_date=datetime(2026, 3, 1, tzinfo=timezone.utc),
        end_date=None,
        description=f"Imported from Inter-company sheet",
        created_at=now,
        updated_at=now,
    )
    await proj.insert()
    return proj


async def _upsert_employee_project(employee_id: str, project_id: str, now: datetime):
    existing = await EmployeeProject.find_one(
        EmployeeProject.employee_id == employee_id,
        EmployeeProject.project_id == project_id,
    )
    if not existing:
        await EmployeeProject(
            employee_id=employee_id,
            project_id=project_id,
            role_in_project="Consultant",
            start_date=datetime(2026, 3, 1, tzinfo=timezone.utc),
            created_at=now,
            updated_at=now,
        ).insert()


async def run(file_path: Path, user_email: str):
    await init_db()

    user = await User.find_one(User.email == user_email)
    if not user:
        raise SystemExit(f"User not found: {user_email}")

    branch_location_id = user.branch_location_id
    print(f"Branch location: {branch_location_id}")

    wb = openpyxl.load_workbook(io.BytesIO(file_path.read_bytes()), data_only=True)
    if TARGET_SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet '{TARGET_SHEET}' not found. Available: {wb.sheetnames}")

    ws = wb[TARGET_SHEET]
    records = _parse_sheet(ws)
    print(f"Parsed {len(records)} YTPL rows with March 2026 hours")

    # Load all active employees for this branch
    employees = await Employee.find(
        Employee.location_id == branch_location_id,
        Employee.is_active == True,
    ).to_list()
    print(f"Loaded {len(employees)} active employees for branch")

    exact_map, norm_map = _build_employee_maps(employees)

    now = datetime.now(timezone.utc)

    # Group by (client, project) to upsert projects once
    project_cache: dict[tuple, Project] = {}
    async def get_project(client: str, project: str) -> Project:
        key = (client, project)
        if key not in project_cache:
            project_cache[key] = await _upsert_project(client, project, now)
        return project_cache[key]

    # Delete existing allocations for this period from inter-company source
    await ProjectAllocation.find(
        ProjectAllocation.period == PERIOD,
        ProjectAllocation.source_system == "intercompany_excel",
    ).delete()

    # --- Pass 1: upsert ALL projects regardless of employee match ---
    for rec in records:
        await get_project(rec["client"], rec["project"])
    print(f"  Projects upserted  : {len(project_cache)}")

    # --- Pass 2: store allocations for matched employees only ---
    matched = 0
    unmatched = []
    allocations_inserted = 0

    for rec in records:
        proj = await get_project(rec["client"], rec["project"])
        proj_id = str(proj.id)

        emp = _resolve_employee(rec["name"], exact_map, norm_map)
        if not emp:
            unmatched.append(rec["name"])
            # Still record a placeholder allocation so the project shows member_count
            # using the raw name so the project is visible even without a DB employee
            hours = rec["hours"]
            allocation_pct = round(min(100.0, hours / CAPACITY_HOURS * 100), 1)
            available_days = round(max(0.0, WORKING_DAYS - hours / HOURS_PER_DAY), 1)
            placeholder_emp_id = f"unmatched:{_normalise_name(rec['name'])}"
            await ProjectAllocation(
                employee_id=placeholder_emp_id,
                hrms_employee_id=0,
                employee_name=rec["name"],
                project_id=proj_id,
                hrms_project_id=getattr(proj, "hrms_project_id", 0) or 0,
                project_name=proj.name,
                client_name=proj.client_name,
                period=PERIOD,
                allocated_days=round(hours / HOURS_PER_DAY, 1),
                allocation_percentage=allocation_pct,
                total_working_days=WORKING_DAYS,
                available_days=available_days,
                source_system="intercompany_excel",
                source_id=f"intercompany:{PERIOD}:{placeholder_emp_id}:{proj_id}",
                created_at=now,
                updated_at=now,
            ).insert()
            allocations_inserted += 1
            continue

        matched += 1
        emp_id = str(emp.id)
        hours = rec["hours"]
        allocation_pct = round(min(100.0, hours / CAPACITY_HOURS * 100), 1)
        available_days = round(max(0.0, WORKING_DAYS - hours / HOURS_PER_DAY), 1)

        await ProjectAllocation(
            employee_id=emp_id,
            hrms_employee_id=getattr(emp, "hrms_employee_id", 0) or 0,
            employee_name=emp.name,
            project_id=proj_id,
            hrms_project_id=getattr(proj, "hrms_project_id", 0) or 0,
            project_name=proj.name,
            client_name=proj.client_name,
            period=PERIOD,
            allocated_days=round(hours / HOURS_PER_DAY, 1),
            allocation_percentage=allocation_pct,
            total_working_days=WORKING_DAYS,
            available_days=available_days,
            source_system="intercompany_excel",
            source_id=f"intercompany:{PERIOD}:{emp_id}:{proj_id}",
            created_at=now,
            updated_at=now,
        ).insert()
        allocations_inserted += 1

        await _upsert_employee_project(emp_id, proj_id, now)

    print(f"\nResults:")
    print(f"  Matched employees  : {matched}")
    print(f"  Unmatched (placeholder allocations): {len(unmatched)}")
    print(f"  Allocations stored : {allocations_inserted}")
    if unmatched:
        print(f"\nUnmatched names (first 20):")
        for n in unmatched[:20]:
            print(f"  - {n}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True)
    parser.add_argument("--user-email", default="manager@nxzen.com")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    asyncio.run(run(path, args.user_email))


if __name__ == "__main__":
    main()
