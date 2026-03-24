"""
Unified Excel seed — imports all relevant sheets from the Utilisation Report workbook.

Sheets processed:
  1. Availability Report - YTPL  → ExcelUtilisationReport (availability fractions per employee per month)
  2. Inter-company               → ProjectAllocation (March 2026 hours per employee per project)
  3. planned days                → EmployeePlannedWorked.planned_days (days per employee per project per month)
  4. Worked days                 → EmployeePlannedWorked.worked_days  (days per employee per project per month)

Usage:
    cd backend
    python -m seed.seed_excel --file "C:/Users/sahit/Downloads/Utilisation_Report_6th_March_2026_updated (2).xlsx" --user-email manager@nxzen.com
"""

import argparse
import asyncio
import io
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.database import init_db
from app.models.employee import Employee
from app.models.employee_planned_worked import EmployeePlannedWorked
from app.models.employee_project import EmployeeProject
from app.models.location import Location
from app.models.project import Project
from app.models.project_allocation import ProjectAllocation
from app.models.user import User
from app.services.excel_utilisation_service import parse_and_store_excel

# ── constants ────────────────────────────────────────────────────────────────
COMPANY_FILTER = {"YTPL"}
IC_SHEET = "Inter-company"
PLANNED_SHEET = "planned days"
WORKED_SHEET = "Worked days"
IC_PERIOD = "2026-03"
IC_WORKING_DAYS = 21
IC_HOURS_PER_DAY = 8.0
IC_CAPACITY_HOURS = IC_WORKING_DAYS * IC_HOURS_PER_DAY

OVERRIDES_FILE = Path(__file__).parent / "match_overrides.json"
HYD_REPORT_SHEET = "Employees HYD"
BGLR_REPORT_SHEET = "Employees BGLR"
UNMATCHED_REPORT_SHEET = "Employees Not Matched"


def _load_overrides() -> dict[str, Optional[str]]:
    """Load manual match overrides: {excel_name -> employee_id or None}."""
    if OVERRIDES_FILE.exists():
        return json.loads(OVERRIDES_FILE.read_text())
    return {}


# ── helpers ───────────────────────────────────────────────────────────────────
def _norm(value: str) -> str:
    c = value.lower().strip()
    c = re.sub(r"\s*[-(]?\s*(onsite|offshore)\s*[)]?\s*$", "", c)
    c = re.sub(r"[^a-z0-9]+", "", c)
    return c


def _meaningful_tokens(name: str) -> set[str]:
    cleaned = name.lower().strip()
    cleaned = re.sub(r"\s*[-(]?\s*(onsite|offshore)\s*[)]?\s*$", "", cleaned)
    return {w for w in re.split(r"[^a-z]+", cleaned) if len(w) >= 3}


def _token_signature(name: str) -> str:
    return " ".join(sorted(_meaningful_tokens(name)))


def _parse_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return max(0.0, float(v))
    except (TypeError, ValueError):
        return 0.0


def _build_emp_maps(employees: list[Employee]):
    exact, normalised, token_index, token_signature_map = {}, {}, {}, {}
    for e in employees:
        exact.setdefault(e.name.strip().lower(), e)
        normalised.setdefault(_norm(e.name), e)
        signature = _token_signature(e.name)
        if signature and signature not in token_signature_map:
            token_signature_map[signature] = e
        for token in _meaningful_tokens(e.name):
            token_index.setdefault(token, []).append(e)
    return exact, normalised, token_index, token_signature_map


def _select_availability_sheet(workbook):
    preferred_names = ["Availability Report - YTPL", "Availability Report - NGL"]
    exact_names = {name.strip().lower(): name for name in workbook.sheetnames}
    for preferred_name in preferred_names:
        actual_name = exact_names.get(preferred_name.strip().lower())
        if actual_name:
            return workbook[actual_name]

    for name in workbook.sheetnames:
        if "ytpl" in name.strip().lower():
            return workbook[name]

    for name in workbook.sheetnames:
        lowered = name.strip().lower()
        if "availability" in lowered and "ngl" not in lowered:
            return workbook[name]

    return None


def _collect_availability_names(workbook) -> list[dict]:
    sheet = _select_availability_sheet(workbook)
    if sheet is None:
        return []

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    month_col_map = {}
    header_row_idx = -1
    for row_index, row in enumerate(rows[:15]):
        for col_index, cell in enumerate(row or []):
            if isinstance(cell, datetime):
                month_col_map[col_index] = f"{cell.year:04d}-{cell.month:02d}"
        if month_col_map:
            header_row_idx = row_index
            break

    if header_row_idx < 0:
        return []

    header = rows[header_row_idx]
    name_col = None
    status_col = None
    for col_index, cell in enumerate(header):
        if col_index in month_col_map:
            continue
        value = str(cell).strip().lower() if cell is not None else ""
        if value == "name":
            name_col = col_index
        elif value == "status":
            status_col = col_index

    if name_col is None:
        name_col = 2
    if status_col is None:
        status_col = 3

    names: list[dict] = []
    for row in rows[header_row_idx + 1:]:
        if not row or len(row) <= name_col:
            continue
        employee_name = str(row[name_col]).strip() if row[name_col] else ""
        if not employee_name or employee_name.lower() in {"name", "total", "grand total"}:
            continue
        if len(row) > status_col and row[status_col]:
            status_value = str(row[status_col]).strip().lower()
            if status_value in {"left", "inactive", "resigned", "terminated"}:
                continue
        names.append({"name": employee_name, "sheet": sheet.title})
    return names


def _collect_workbook_employee_sources(workbook) -> dict[str, dict]:
    employee_sources: dict[str, dict] = {}

    def add_name(employee_name: str, sheet_name: str):
        clean_name = employee_name.strip()
        if not clean_name:
            return
        key = _norm(clean_name)
        if not key:
            return
        bucket = employee_sources.setdefault(
            key,
            {
                "excel_name": clean_name,
                "sheet_names": set(),
            },
        )
        bucket["sheet_names"].add(sheet_name)

    for item in _collect_availability_names(workbook):
        add_name(item["name"], item["sheet"])

    if IC_SHEET in workbook.sheetnames:
        for rec in _parse_intercompany(workbook[IC_SHEET]):
            add_name(rec["name"], IC_SHEET)

    if PLANNED_SHEET in workbook.sheetnames:
        for rec in _parse_days_sheet(workbook[PLANNED_SHEET]):
            add_name(rec["name"], PLANNED_SHEET)

    if WORKED_SHEET in workbook.sheetnames:
        for rec in _parse_days_sheet(workbook[WORKED_SHEET]):
            add_name(rec["name"], WORKED_SHEET)

    return employee_sources


def _location_bucket(location: Optional[Location]) -> str:
    if not location:
        return "unmatched"

    code = (location.code or "").strip().upper()
    city = (location.city or "").strip().lower()
    if code == "HYD" or "hyderabad" in city:
        return "hyd"
    if code in {"BLR", "BGLR"} or city in {"bangalore", "bengaluru"}:
        return "bglr"
    return "unmatched"


def _write_category_sheet(workbook, title: str, rows: list[dict]):
    if title in workbook.sheetnames:
        del workbook[title]

    ws = workbook.create_sheet(title)
    headers = [
        "Excel Name",
        "Matched HRMS Name",
        "HRMS Employee ID",
        "Location Code",
        "Location City",
        "Match Status",
        "Source Sheets",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(
            [
                row["excel_name"],
                row.get("hrms_name") or "",
                row.get("hrms_employee_id") or "",
                row.get("location_code") or "",
                row.get("location_city") or "",
                row["match_status"],
                row["source_sheets"],
            ]
        )

    for column_index, _ in enumerate(headers, start=1):
        max_len = 0
        for cell in ws[get_column_letter(column_index)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 14), 42)

    ws.freeze_panes = "A2"


async def update_workbook_employee_split(file_path: Path) -> Path:
    report_workbook = openpyxl.load_workbook(file_path)
    data_workbook = openpyxl.load_workbook(file_path, data_only=True)

    employees = await Employee.find(Employee.is_active == True).to_list()
    locations = await Location.find_all().to_list()
    location_by_id = {str(location.id): location for location in locations}
    exact_map, norm_map, token_map, token_signature_map = _build_emp_maps(employees)
    emp_by_id = {str(employee.id): employee for employee in employees}
    overrides = _load_overrides()

    employee_sources = _collect_workbook_employee_sources(data_workbook)
    buckets = {
        "hyd": [],
        "bglr": [],
        "unmatched": [],
    }

    for item in employee_sources.values():
        excel_name = item["excel_name"]
        matched_employee = _resolve(
            excel_name,
            exact_map,
            norm_map,
            token_map,
            token_signature_map,
            overrides,
            emp_by_id,
        )
        location = location_by_id.get(matched_employee.location_id) if matched_employee and matched_employee.location_id else None
        bucket_key = _location_bucket(location)
        buckets[bucket_key].append(
            {
                "excel_name": excel_name,
                "hrms_name": matched_employee.name if matched_employee else "",
                "hrms_employee_id": getattr(matched_employee, "hrms_employee_id", "") if matched_employee else "",
                "location_code": location.code if location else "",
                "location_city": location.city if location else "",
                "match_status": "Matched in HRMS" if matched_employee else "Not matched / missing in HRMS",
                "source_sheets": ", ".join(sorted(item["sheet_names"])),
            }
        )

    for bucket_rows in buckets.values():
        bucket_rows.sort(key=lambda row: row["excel_name"].lower())

    _write_category_sheet(report_workbook, HYD_REPORT_SHEET, buckets["hyd"])
    _write_category_sheet(report_workbook, BGLR_REPORT_SHEET, buckets["bglr"])
    _write_category_sheet(report_workbook, UNMATCHED_REPORT_SHEET, buckets["unmatched"])
    report_workbook.save(file_path)
    return file_path


def _fuzzy_match(name: str, token_index: dict, threshold: float = 0.6) -> Optional[Employee]:
    query_tokens = _meaningful_tokens(name)
    if not query_tokens:
        return None
    candidates: dict[str, tuple] = {}
    for token in query_tokens:
        for emp in token_index.get(token, []):
            eid = str(emp.id)
            if eid not in candidates:
                candidates[eid] = (emp, 0)
            candidates[eid] = (emp, candidates[eid][1] + 1)
    best_emp, best_score = None, 0.0
    for emp, common in candidates.values():
        emp_tokens = _meaningful_tokens(emp.name)
        denom = max(len(query_tokens), len(emp_tokens))
        score = common / denom if denom else 0.0
        if score > best_score:
            best_score, best_emp = score, emp
    return best_emp if best_score >= threshold else None


def _resolve_override(name: str, overrides: dict, emp_by_id: Optional[dict[str, Employee]]) -> tuple[bool, Optional[Employee]]:
    if name not in overrides:
        return False, None
    override_id = overrides[name]
    if override_id is None:
        return True, None
    if emp_by_id and override_id in emp_by_id:
        return True, emp_by_id[override_id]
    return False, None


def _unique_token_candidate(name: str, token_index: dict) -> Optional[Employee]:
    query_tokens = _meaningful_tokens(name)
    if not query_tokens:
        return None
    candidate_ids: set[str] = set()
    candidate_map: dict[str, Employee] = {}
    for token in query_tokens:
        for emp in token_index.get(token, []):
            emp_id = str(emp.id)
            candidate_ids.add(emp_id)
            candidate_map[emp_id] = emp
    if len(candidate_ids) == 1:
        candidate = candidate_map[next(iter(candidate_ids))]
        candidate_tokens = _meaningful_tokens(candidate.name)
        if query_tokens.issubset(candidate_tokens):
            return candidate
    return None


def _resolve(name: str, exact, normalised, token_index=None, token_signature_map=None, overrides: dict = {}, emp_by_id: Optional[dict[str, Employee]] = None) -> Optional[Employee]:
    # Manual override takes priority
    override_hit, override_employee = _resolve_override(name, overrides, emp_by_id)
    if override_hit:
        return override_employee
        # We need the Employee object — caller must handle by ID lookup
        # Return a sentinel so caller knows to fetch by ID
    result = exact.get(name.strip().lower()) or normalised.get(_norm(name))
    if result is None and token_signature_map is not None:
        result = token_signature_map.get(_token_signature(name))
    if result is None and token_index is not None:
        result = _fuzzy_match(name, token_index, threshold=0.6)
    if result is None and token_index is not None:
        result = _unique_token_candidate(name, token_index)
    return result


# ── sheet 2: inter-company ────────────────────────────────────────────────────
def _parse_intercompany(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = march_col = None
    for idx, row in enumerate(rows[:10]):
        for i, cell in enumerate(row or []):
            if isinstance(cell, datetime) and cell.year == 2026 and cell.month == 3:
                header_idx, march_col = idx, i
                break
        if march_col is not None:
            break
    if march_col is None:
        raise ValueError("March 2026 column not found in Inter-company sheet")

    records = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) <= march_col:
            continue
        name = str(row[1]).strip() if row[1] else ""
        company = str(row[2]).strip() if row[2] else ""
        client = str(row[4]).strip() if row[4] else ""
        project = str(row[6]).strip() if row[6] else ""
        if not name or company.upper() not in COMPANY_FILTER:
            continue
        nl = name.lower().strip()
        if nl.startswith(("tbc", "total", "sub-total")):
            continue
        hours = _parse_float(row[march_col])
        if hours <= 0:
            continue
        records.append({"name": name, "client": client, "project": project, "hours": hours})
    return records


async def _upsert_project(client_name: str, project_name: str, now: datetime) -> Project:
    existing = await Project.find_one(
        Project.name == project_name, Project.client_name == client_name, Project.is_deleted != True
    ) or await Project.find_one(Project.name == project_name, Project.is_deleted != True)
    if existing:
        if not existing.client_name:
            existing.client_name = client_name
            existing.updated_at = now
            await existing.save()
        return existing
    p = Project(
        name=project_name, client_name=client_name, status="ACTIVE", project_type="client",
        start_date=datetime(2026, 3, 1, tzinfo=timezone.utc), created_at=now, updated_at=now,
        description="Imported from Inter-company sheet",
    )
    await p.insert()
    return p


async def import_intercompany(wb, branch_location_id: str, now: datetime) -> dict:
    if IC_SHEET not in wb.sheetnames:
        return {"skipped": True, "reason": f"Sheet '{IC_SHEET}' not found"}

    records = _parse_intercompany(wb[IC_SHEET])
    print(f"  [Inter-company] Parsed {len(records)} YTPL rows with March 2026 hours")

    employees = await Employee.find(Employee.is_active == True).to_list()
    exact_map, norm_map, token_map, token_signature_map = _build_emp_maps(employees)
    emp_by_id = {str(e.id): e for e in employees}
    overrides = _load_overrides()

    project_cache: dict[tuple, Project] = {}

    async def get_proj(client: str, proj_name: str) -> Project:
        key = (client, proj_name)
        if key not in project_cache:
            project_cache[key] = await _upsert_project(client, proj_name, now)
        return project_cache[key]

    # Delete existing
    await ProjectAllocation.find(
        ProjectAllocation.period == IC_PERIOD,
        ProjectAllocation.source_system == "intercompany_excel",
    ).delete()

    # Pass 1: upsert all projects
    for rec in records:
        await get_proj(rec["client"], rec["project"])

    # Pass 2: allocations
    matched = unmatched = inserted = skipped_other_branch = 0
    for rec in records:
        proj = await get_proj(rec["client"], rec["project"])
        proj_id = str(proj.id)
        emp = _resolve(rec["name"], exact_map, norm_map, token_map, token_signature_map, overrides, emp_by_id)
        if emp and emp.location_id != branch_location_id:
            skipped_other_branch += 1
            continue

        hours = rec["hours"]
        alloc_pct = round(min(100.0, hours / IC_CAPACITY_HOURS * 100), 1)
        avail_days = round(max(0.0, IC_WORKING_DAYS - hours / IC_HOURS_PER_DAY), 1)

        if emp:
            emp_id = str(emp.id)
            emp_name = emp.name
            hrms_id = getattr(emp, "hrms_employee_id", 0) or 0
            matched += 1
        else:
            emp_id = f"unmatched:{_norm(rec['name'])}"
            emp_name = rec["name"]
            hrms_id = 0
            unmatched += 1

        await ProjectAllocation(
            employee_id=emp_id, hrms_employee_id=hrms_id, employee_name=emp_name,
            project_id=proj_id, hrms_project_id=getattr(proj, "hrms_project_id", 0) or 0,
            project_name=proj.name, client_name=proj.client_name, period=IC_PERIOD,
            allocated_days=round(hours / IC_HOURS_PER_DAY, 1), allocation_percentage=alloc_pct,
            total_working_days=IC_WORKING_DAYS, available_days=avail_days,
            source_system="intercompany_excel",
            source_id=f"intercompany:{IC_PERIOD}:{emp_id}:{proj_id}",
            created_at=now, updated_at=now,
        ).insert()
        inserted += 1

        if emp:
            if not await EmployeeProject.find_one(
                EmployeeProject.employee_id == emp_id, EmployeeProject.project_id == proj_id
            ):
                await EmployeeProject(
                    employee_id=emp_id, project_id=proj_id, role_in_project="Consultant",
                    start_date=datetime(2026, 3, 1, tzinfo=timezone.utc), created_at=now, updated_at=now,
                ).insert()

    print(f"  [Inter-company] Projects: {len(project_cache)}, Allocations: {inserted} (matched: {matched}, unmatched: {unmatched}, skipped_other_branch: {skipped_other_branch})")
    return {"projects": len(project_cache), "allocations": inserted, "matched": matched, "unmatched": unmatched, "skipped_other_branch": skipped_other_branch}


# ── sheets 3 & 4: planned days / worked days ──────────────────────────────────
def _parse_days_sheet(ws) -> list[dict]:
    """
    Parse a 'planned days' or 'Worked days' sheet.
    Returns list of {name, company, client, project, period -> days}.
    Header row 0: Name, Company Name, Band, Account, Project Name (Revenue),
                  Project Name (Commercial), Location, <datetime months...>
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    # Build period map: col_index -> "YYYY-MM"
    period_cols: dict[int, str] = {}
    for i, cell in enumerate(header):
        if isinstance(cell, datetime):
            period_cols[i] = f"{cell.year:04d}-{cell.month:02d}"

    records = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        company = str(row[1]).strip() if row[1] else ""
        client = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        project = str(row[5]).strip() if len(row) > 5 and row[5] else ""  # Commercial name

        if not name or company.upper() not in COMPANY_FILTER:
            continue
        nl = name.lower()
        if nl.startswith(("tbc", "total", "sub-total", "name")):
            continue

        month_data: dict[str, float] = {}
        for col_idx, period in period_cols.items():
            if col_idx < len(row):
                val = _parse_float(row[col_idx])
                if val > 0:
                    month_data[period] = val

        if month_data:
            records.append({
                "name": name,
                "client": client,
                "project": project,
                "months": month_data,
            })
    return records


async def import_planned_worked(wb, branch_location_id: str, filename: str, now: datetime) -> dict:
    has_planned = PLANNED_SHEET in wb.sheetnames
    has_worked = WORKED_SHEET in wb.sheetnames

    if not has_planned and not has_worked:
        return {"skipped": True, "reason": "Neither 'planned days' nor 'Worked days' sheets found"}

    employees = await Employee.find(Employee.is_active == True).to_list()
    exact_map, norm_map, token_map, token_signature_map = _build_emp_maps(employees)
    emp_by_id = {str(e.id): e for e in employees}
    overrides = _load_overrides()

    # Parse both sheets
    planned_records = _parse_days_sheet(wb[PLANNED_SHEET]) if has_planned else []
    worked_records = _parse_days_sheet(wb[WORKED_SHEET]) if has_worked else []

    print(f"  [Planned/Worked] Planned rows: {len(planned_records)}, Worked rows: {len(worked_records)}")

    # Build lookup: (norm_name, period, norm_project) -> days
    planned_lookup: dict[tuple, float] = {}
    for rec in planned_records:
        key_name = _norm(rec["name"])
        key_proj = _norm(rec["project"])
        for period, days in rec["months"].items():
            planned_lookup[(key_name, period, key_proj)] = days

    worked_lookup: dict[tuple, float] = {}
    for rec in worked_records:
        key_name = _norm(rec["name"])
        key_proj = _norm(rec["project"])
        for period, days in rec["months"].items():
            worked_lookup[(key_name, period, key_proj)] = days

    # Collect all unique (name, period, project, client) combos
    all_keys: set[tuple] = set()
    for rec in planned_records:
        for period in rec["months"]:
            all_keys.add((_norm(rec["name"]), period, _norm(rec["project"]), rec["name"], rec["project"], rec["client"]))
    for rec in worked_records:
        for period in rec["months"]:
            all_keys.add((_norm(rec["name"]), period, _norm(rec["project"]), rec["name"], rec["project"], rec["client"]))

    # Delete existing data for this source
    await EmployeePlannedWorked.find(
        {"source_file": filename}
    ).delete()

    inserted = matched = unmatched = skipped_other_branch = 0
    for norm_name, period, norm_proj, raw_name, raw_proj, client in all_keys:
        emp = _resolve(raw_name, exact_map, norm_map, token_map, token_signature_map, overrides, emp_by_id)
        if emp and emp.location_id != branch_location_id:
            skipped_other_branch += 1
            continue
        emp_id = str(emp.id) if emp else f"unmatched:{norm_name}"
        emp_display = emp.name if emp else raw_name
        if emp:
            matched += 1
        else:
            unmatched += 1

        planned = planned_lookup.get((norm_name, period, norm_proj), 0.0)
        worked = worked_lookup.get((norm_name, period, norm_proj), 0.0)

        try:
            await EmployeePlannedWorked(
                employee_id=emp_id,
                employee_name=emp_display,
                project_name=raw_proj,
                client_name=client or None,
                period=period,
                planned_days=planned,
                worked_days=worked,
                source_file=filename,
                imported_at=now,
            ).insert()
            inserted += 1
        except Exception:
            # Duplicate — update instead
            existing = await EmployeePlannedWorked.find_one(
                EmployeePlannedWorked.employee_id == emp_id,
                EmployeePlannedWorked.period == period,
                EmployeePlannedWorked.project_name == raw_proj,
            )
            if existing:
                existing.planned_days = planned
                existing.worked_days = worked
                existing.imported_at = now
                await existing.save()

    print(f"  [Planned/Worked] Inserted: {inserted}, Matched employees: {matched}, Unmatched: {unmatched}, Skipped other branch: {skipped_other_branch}")
    return {"inserted": inserted, "matched": matched, "unmatched": unmatched, "skipped_other_branch": skipped_other_branch}


# ── main ──────────────────────────────────────────────────────────────────────
async def run(file_path: Path, user_email: str):
    await init_db()

    user = await User.find_one(User.email == user_email)
    if not user:
        raise SystemExit(f"User not found: {user_email}")

    branch_location_id = user.branch_location_id
    print(f"Branch: {branch_location_id}")

    content = file_path.read_bytes()
    filename = file_path.name
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    now = datetime.now(timezone.utc)

    print("\n[1/3] Importing YTPL availability sheet...")
    ytpl_result = await parse_and_store_excel(
        file_content=content,
        filename=filename,
        branch_location_id=branch_location_id,
        user_id=str(user.id),
    )
    print(f"  Rows stored: {ytpl_result.get('rows_stored', 0)}, Periods: {ytpl_result.get('periods', [])}")

    print("\n[2/3] Importing Inter-company sheet...")
    ic_result = await import_intercompany(wb, branch_location_id, now)

    print("\n[3/3] Importing Planned/Worked days sheets...")
    pw_result = await import_planned_worked(wb, branch_location_id, filename, now)

    print("\n[4/4] Updating workbook with employee split sheets...")
    updated_path = await update_workbook_employee_split(file_path)

    print("\n=== Done ===")
    print(f"  YTPL availability : {ytpl_result.get('rows_stored', 0)} rows")
    print(f"  Inter-company     : {ic_result.get('allocations', 0)} allocations across {ic_result.get('projects', 0)} projects")
    print(f"  Planned/Worked    : {pw_result.get('inserted', 0)} records")
    print(f"  Workbook updated  : {updated_path}")


def main():
    parser = argparse.ArgumentParser(description="Unified Excel seed for the Utilisation Report workbook.")
    parser.add_argument("--file", required=True, help="Path to the .xlsx workbook")
    parser.add_argument("--user-email", default="manager@nxzen.com")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    asyncio.run(run(path, args.user_email))


if __name__ == "__main__":
    main()
